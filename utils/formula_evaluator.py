"""
formula_evaluator.py
--------------------
使用 formulas 套件評估 Excel 公式，解決 openpyxl data_only=True
在寫入後清空公式快取導致公式格回傳 None 的問題。

支援：
- 基本算術公式（=C12, =SUM(...), =A1*B1 等）
- 跨 sheet 引用（=公版!C16 等）
- 動態長度（橫向讀取年度現金流）

若 formulas 未安裝，自動降級為 openpyxl data_only=True。
"""

import os
from typing import Any, List, Optional

import numpy as np


try:
    import formulas as _formulas_module
    _FORMULAS_OK = True
except ImportError:
    _FORMULAS_OK = False
    print("[FormulaEvaluator] 警告: formulas 套件未安裝，請執行 pip install formulas")
    print("[FormulaEvaluator] 降級為 openpyxl data_only=True（公式格可能讀到 None）")

# 模組級快取：同一檔案在同一 mtime 內只計算一次
# key = (abs_path, mtime)，value = FormulaEvaluator 實例
_evaluator_cache: dict = {}


def get_evaluator(file_path: str) -> "FormulaEvaluator":
    """取得（或重用）指定檔案的 FormulaEvaluator，檔案變更後自動失效。"""
    abs_path = os.path.abspath(file_path)
    try:
        mtime = os.path.getmtime(abs_path)
    except OSError:
        mtime = None
    key = (abs_path, mtime)
    if key not in _evaluator_cache:
        _evaluator_cache.clear()   # 清除舊快取，避免記憶體無限增長
        _evaluator_cache[key] = FormulaEvaluator(abs_path)
    return _evaluator_cache[key]


class FormulaEvaluator:
    """
    封裝 formulas 套件，提供統一的儲存格評估介面。
    同一個 FormulaEvaluator 實例代表一個已載入並計算完畢的 workbook，
    可重複呼叫 get_cell() 而不重複計算。
    """

    def __init__(self, file_path: str):
        self.file_path = os.path.abspath(file_path)
        self._model = None          # formulas ExcelModel（None 表示尚未載入）
        self._load_failed = False   # 標記 formulas 載入是否失敗

    # ------------------------------------------------------------------
    # 公開 API
    # ------------------------------------------------------------------

    def get_cell(self, sheet_name: str, cell_ref: str) -> Optional[Any]:
        """
        取得指定儲存格的評估後數值。
        公式會被 formulas 套件正確計算（不依賴 Excel 快取）。

        Args:
            sheet_name: 工作表名稱，如「公版」、「滾算紀錄1」
            cell_ref:   儲存格位址，如「C16」、「B5」

        Returns:
            評估後的數值（int / float / str / None）
        """
        self._ensure_loaded()

        if not self._load_failed and self._model is not None:
            val = self._get_from_formulas(sheet_name, cell_ref.upper())
            if val is not None:
                return val
            # formulas 找不到此格（可能是空格）→ 也嘗試 openpyxl
        return self._get_from_openpyxl(sheet_name, cell_ref.upper())

    def get_row_values(
        self,
        sheet_name: str,
        row: int,
        col_start: int,
        col_end: int
    ) -> List[Optional[Any]]:
        """
        橫向讀取一整行的評估值（用於讀取現金流序列）。

        Args:
            sheet_name: 工作表名稱
            row:        行號（1-based）
            col_start:  起始欄號（1-based，C=3）
            col_end:    結束欄號（1-based）

        Returns:
            按欄位順序排列的值清單
        """
        from openpyxl.utils import get_column_letter
        return [
            self.get_cell(sheet_name, f"{get_column_letter(col)}{row}")
            for col in range(col_start, col_end + 1)
        ]

    # ------------------------------------------------------------------
    # 內部方法
    # ------------------------------------------------------------------

    def _ensure_loaded(self):
        """第一次呼叫時載入並計算 workbook，之後快取結果。"""
        if self._model is not None or self._load_failed:
            return
        if not _FORMULAS_OK:
            print(f"[FormulaEvaluator] ❌ formulas 套件未安裝，無法評估公式")
            self._load_failed = True
            return
        print(f"[FormulaEvaluator] 正在載入: {self.file_path}")
        try:
            xl = _formulas_module.ExcelModel().loads(self.file_path).finish()
            # 先把 model 存起來，即使 calculate() 部分失敗仍可讀已計算的 cell
            self._model = xl
            try:
                xl.calculate()
                print(f"[FormulaEvaluator] ✅ 載入成功，共 {len(xl.cells)} 個儲存格")
            except Exception as calc_e:
                print(f"[FormulaEvaluator] ⚠️ 部分公式計算失敗（可能含不支援的函數如 IRR）: {calc_e}")
                print(f"[FormulaEvaluator] 繼續使用已計算的儲存格（其餘 cell 降級為 openpyxl）")
        except Exception as e:
            print(f"[FormulaEvaluator] ❌ formulas 載入失敗，降級為 openpyxl: {e}")
            import traceback
            traceback.print_exc()
            self._load_failed = True

    def _get_from_formulas(self, sheet_name: str, cell_ref: str) -> Optional[Any]:
        """從 formulas 模型取得已評估的儲存格值。"""
        fname = os.path.basename(self.file_path)

        # formulas 的 cell key 格式：'[filename.xlsx]SheetName'!REF
        # 嘗試大小寫的不同組合，提高相容性
        candidate_keys = [
            f"'[{fname}]{sheet_name}'!{cell_ref}",
            f"'[{fname.upper()}]{sheet_name}'!{cell_ref}",
            f"'[{fname.lower()}]{sheet_name}'!{cell_ref}",
        ]

        for key in candidate_keys:
            cell = self._model.cells.get(key)
            if cell is not None:
                return self._to_scalar(cell.value)

        # 最後手段：用後綴模糊比對（處理路徑格式不一致）
        suffix_variants = [
            f"]{sheet_name}'!{cell_ref}",
            f"]{sheet_name}!{cell_ref}",
        ]
        for key, cell in self._model.cells.items():
            key_upper = key.upper()
            for suffix in suffix_variants:
                if key_upper.endswith(suffix.upper()):
                    return self._to_scalar(cell.value)

        # 找不到時印出可用 key 的前幾筆供診斷
        sample_keys = list(self._model.cells.keys())[:5]
        print(f"[FormulaEvaluator] ⚠️ 找不到 [{sheet_name}]{cell_ref}，model key 範例: {sample_keys}")
        return None  # formulas 中找不到此格

    def _get_from_openpyxl(self, sheet_name: str, cell_ref: str) -> Optional[Any]:
        """降級方案：用 openpyxl data_only=True 讀取（公式快取已清空時可能回傳 None）。"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            if sheet_name in wb.sheetnames:
                val = wb[sheet_name][cell_ref].value
                wb.close()
                return val
            wb.close()
        except Exception:
            pass
        return None

    @staticmethod
    def _to_scalar(val: Any) -> Optional[Any]:
        """將 formulas 回傳的 numpy array 或巢狀結構轉換為單一純量值。"""
        if val is None:
            return None
        # numpy array（formulas 通常回傳 shape=(1,1) 的 array）
        if isinstance(val, np.ndarray):
            flat = val.flat
            try:
                v = next(iter(flat))
                return None if (isinstance(v, float) and np.isnan(v)) else v
            except StopIteration:
                return None
        # 巢狀 list/tuple（部分版本會回傳 [[value]]）
        if isinstance(val, (list, tuple)):
            try:
                inner = val[0]
                if isinstance(inner, (list, tuple)):
                    inner = inner[0]
                return inner
            except (IndexError, TypeError):
                return None
        return val
