"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice
Supports Windows, macOS, and Linux
"""

import json
import os
import platform
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

MACRO_DIR_MACOS   = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
MACRO_DIR_LINUX   = "~/.config/libreoffice/4/user/basic/Standard"
MACRO_DIR_WINDOWS = os.path.join(os.environ.get("APPDATA", ""), "LibreOffice", "4", "user", "basic", "Standard")
MACRO_FILENAME    = "Module1.xba"

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

# Windows 上 LibreOffice soffice.exe 的常見安裝路徑
WINDOWS_SOFFICE_PATHS = [
    r"D:\LibreOffice\program\soffice.exe",
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
]


def get_soffice_cmd():
    """取得 soffice 執行檔路徑（跨平台）"""
    if platform.system() == "Windows":
        # 優先用 PATH 裡的 soffice
        import shutil
        found = shutil.which("soffice")
        if found:
            return found
        # 找不到就嘗試預設安裝路徑
        for p in WINDOWS_SOFFICE_PATHS:
            if os.path.exists(p):
                return p
        raise FileNotFoundError(
            "找不到 LibreOffice。請確認已安裝，並將 "
            r"C:\Program Files\LibreOffice\program 加入系統 PATH。"
        )
    return "soffice"


def get_soffice_env():
    """取得執行 soffice 所需的環境變數（Windows 不需要 LD_PRELOAD）"""
    env = os.environ.copy()
    if platform.system() != "Windows":
        env["SAL_USE_VCLPLUGIN"] = "svp"
    return env


def setup_libreoffice_macro():
    """確保 LibreOffice macro 已安裝"""
    system = platform.system()
    if system == "Darwin":
        macro_dir = os.path.expanduser(MACRO_DIR_MACOS)
    elif system == "Windows":
        macro_dir = MACRO_DIR_WINDOWS
    else:
        macro_dir = os.path.expanduser(MACRO_DIR_LINUX)

    macro_file = os.path.join(macro_dir, MACRO_FILENAME)

    if (
        os.path.exists(macro_file)
        and "RecalculateAndSave" in Path(macro_file).read_text(encoding="utf-8")
    ):
        return True

    # macro 不存在，先讓 LibreOffice 初始化建立資料夾
    soffice = get_soffice_cmd()
    if not os.path.exists(macro_dir):
        subprocess.run(
            [soffice, "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=30,
            env=get_soffice_env(),
        )
        os.makedirs(macro_dir, exist_ok=True)

    try:
        Path(macro_file).write_text(RECALCULATE_MACRO, encoding="utf-8")
        return True
    except Exception as e:
        print(f"[recalc] macro 寫入失敗: {e}", file=sys.stderr)
        return False


def recalc(filename, timeout=60):
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).absolute())

    # Windows 路徑需轉換為 file:// URI 格式給 LibreOffice
    if platform.system() == "Windows":
        abs_path_uri = Path(filename).absolute().as_uri()
    else:
        abs_path_uri = abs_path

    if not setup_libreoffice_macro():
        return {"error": "Failed to setup LibreOffice macro"}

    soffice = get_soffice_cmd()

    cmd = [
        soffice,
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path_uri,
    ]

    # timeout 包裝（Linux / macOS 用系統指令，Windows 用 subprocess timeout）
    if platform.system() == "Linux":
        cmd = ["timeout", str(timeout)] + cmd
    elif platform.system() == "Darwin":
        try:
            subprocess.run(["gtimeout", "--version"], capture_output=True, timeout=1)
            cmd = ["gtimeout", str(timeout)] + cmd
        except (FileNotFoundError, subprocess.TimeoutExpired):
            pass

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            env=get_soffice_env(),
            timeout=timeout + 10,
        )
    except subprocess.TimeoutExpired:
        return {"error": f"LibreOffice 超時（>{timeout}s）"}
    except FileNotFoundError as e:
        return {"error": str(e)}

    if result.returncode not in (0, 124):
        error_msg = result.stderr or result.stdout or "Unknown error"
        return {"error": f"LibreOffice returncode={result.returncode}: {error_msg}"}

    # 掃描公式錯誤
    try:
        wb = load_workbook(filename, data_only=True)
        excel_errors = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                error_details[err].append(f"{sheet_name}!{cell.coordinate}")
                                total_errors += 1
                                break
        wb.close()

        output = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": {},
        }
        for err_type, locations in error_details.items():
            if locations:
                output["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],
                }

        # 計算公式數量
        wb2 = load_workbook(filename, data_only=False)
        formula_count = sum(
            1
            for sn in wb2.sheetnames
            for row in wb2[sn].iter_rows()
            for cell in row
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("=")
        )
        wb2.close()
        output["total_formulas"] = formula_count

        return output

    except Exception as e:
        return {"error": str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        sys.exit(1)

    filename = sys.argv[1]
    timeout  = int(sys.argv[2]) if len(sys.argv) > 2 else 60
    result   = recalc(filename, timeout)
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
