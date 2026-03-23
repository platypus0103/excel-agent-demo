# tool/excel_tool.py
import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Optional, Union, List, Tuple
from difflib import SequenceMatcher
from utils.recalc import recalc as _recalc

class ExcelTool:
    """Excel 操作工具類"""

    def __init__(self, file_path: str = "Excel/data.xlsx"):
        """
        初始化 Excel 工具

        Args:
            file_path: Excel 文件路徑（相對於專案根目錄）
        """
        self.file_path = file_path
        
        # 如果使用預設路徑且檔案不存在，嘗試搜尋現有檔案
        if self.file_path == "Excel/data.xlsx" and not os.path.exists(self.file_path):
            existing_file = self._find_existing_file()
            if existing_file:
                self.file_path = existing_file
                print(f"自動選用現有 Excel 檔案: {self.file_path}")

        self._ensure_file_exists()

    def _find_existing_file(self) -> Optional[str]:
        """搜尋 Excel 資料夾中現有的 Excel 檔案"""
        directory = os.path.dirname(self.file_path)
        if os.path.exists(directory):
            for file in os.listdir(directory):
                if file.endswith(('.xlsx', '.xls')) and not file.startswith('~$'):
                    return os.path.join(directory, file)
        return None

    def _ensure_file_exists(self):
        """確保 Excel 文件存在"""
        # 確保目錄存在
        directory = os.path.dirname(self.file_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory)

        # 如果文件不存在，創建新文件
        if not os.path.exists(self.file_path):
            wb = Workbook()
            wb.save(self.file_path)
            print(f"創建新的 Excel 文件: {self.file_path}")

    def _parse_cell_range(self, cell_range: str) -> tuple:
        """
        解析儲存格範圍

        Args:
            cell_range: 儲存格範圍，如 "A1" 或 "A1:B5"

        Returns:
            (start_cell, end_cell) 如果是範圍，否則 (cell, None)
        """
        if ':' in cell_range:
            start, end = cell_range.split(':')
            return start.strip(), end.strip()
        return cell_range.strip(), None

    def write_cell(self, cell: str, value: Union[str, int, float]) -> dict:
        """
        寫入單個儲存格或範圍

        Args:
            cell: 儲存格位置，如 "A1" 或範圍 "A1:B2"
            value: 要寫入的值

        Returns:
            操作結果字典
        """
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active

            start_cell, end_cell = self._parse_cell_range(cell)

            if end_cell:
                # 範圍操作：將相同的值寫入所有儲存格
                for row in ws[start_cell:end_cell]:
                    for cell_obj in row:
                        cell_obj.value = value
                message = f"已將 {value} 寫入範圍 {cell}"
            else:
                # 單個儲存格操作
                ws[start_cell] = value
                message = f"已將 {value} 寫入儲存格 {cell}"

            wb.save(self.file_path)
            _recalc(self.file_path)
            return {
                "success": True,
                "message": message,
                "cell": cell,
                "value": value
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"寫入失敗: {str(e)}",
                "cell": cell,
                "value": value
            }

    def read_cell(self, cell: str) -> dict:
        """
        讀取儲存格或範圍

        Args:
            cell: 儲存格位置，如 "A1" 或範圍 "A1:B2"

        Returns:
            操作結果字典
        """
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active

            start_cell, end_cell = self._parse_cell_range(cell)

            if end_cell:
                # 範圍操作：讀取所有儲存格的值
                values = []
                for row in ws[start_cell:end_cell]:
                    row_values = [cell_obj.value for cell_obj in row]
                    values.append(row_values)
                return {
                    "success": True,
                    "message": f"已讀取範圍 {cell}",
                    "cell": cell,
                    "value": values
                }
            else:
                # 單個儲存格操作
                value = ws[start_cell].value
                return {
                    "success": True,
                    "message": f"已讀取儲存格 {cell}",
                    "cell": cell,
                    "value": value
                }
        except Exception as e:
            return {
                "success": False,
                "message": f"讀取失敗: {str(e)}",
                "cell": cell
            }

    def delete_cell(self, cell: str) -> dict:
        """
        清空儲存格或範圍的內容

        Args:
            cell: 儲存格位置，如 "A1" 或範圍 "A1:B2"

        Returns:
            操作結果字典
        """
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active

            start_cell, end_cell = self._parse_cell_range(cell)

            if end_cell:
                # 範圍操作：清空所有儲存格
                for row in ws[start_cell:end_cell]:
                    for cell_obj in row:
                        cell_obj.value = None
                message = f"已清空範圍 {cell}"
            else:
                # 單個儲存格操作
                ws[start_cell].value = None
                message = f"已清空儲存格 {cell}"

            wb.save(self.file_path)
            _recalc(self.file_path)
            return {
                "success": True,
                "message": message,
                "cell": cell
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"刪除失敗: {str(e)}",
                "cell": cell
            }

    # ========== 智慧編輯功能 ==========

    def _fuzzy_match(self, keyword: str, candidates: List[Tuple[int, str]]) -> Optional[Tuple[int, str, float]]:
        """
        模糊匹配欄位名稱

        Args:
            keyword: 使用者輸入的關鍵字
            candidates: [(row_number, field_name), ...] 候選欄位列表

        Returns:
            (row_number, matched_field_name, similarity_score) 或 None
        """
        keyword_clean = keyword.strip().lower()
        best_match = None
        best_score = 0.0

        for row_num, field_name in candidates:
            # 移除 "Inflow : " 或 "Outflow : " 前綴後比對
            field_clean = field_name.strip()
            field_for_match = field_clean.lower()

            # 移除前綴
            for prefix in ['inflow :', 'outflow :', 'inflow:', 'outflow:']:
                if field_for_match.startswith(prefix):
                    field_for_match = field_for_match[len(prefix):].strip()
                    break

            # 計算相似度
            # 1. 完全匹配（忽略大小寫）
            if keyword_clean == field_for_match:
                return (row_num, field_clean, 1.0)

            # 2. 關鍵字包含在欄位名稱中
            if keyword_clean in field_for_match:
                score = len(keyword_clean) / len(field_for_match) + 0.5
                if score > best_score:
                    best_score = score
                    best_match = (row_num, field_clean, score)
                continue

            # 3. 欄位名稱包含在關鍵字中
            if field_for_match in keyword_clean:
                score = len(field_for_match) / len(keyword_clean) + 0.3
                if score > best_score:
                    best_score = score
                    best_match = (row_num, field_clean, score)
                continue

            # 4. 使用序列匹配器計算相似度（處理錯字）
            similarity = SequenceMatcher(None, keyword_clean, field_for_match).ratio()
            if similarity > best_score and similarity > 0.5:  # 至少 50% 相似
                best_score = similarity
                best_match = (row_num, field_clean, similarity)

        return best_match

    def _parse_year_spec(self, year_spec: str, available_years: List[int]) -> List[int]:
        """
        解析年份規格

        Args:
            year_spec: 年份規格字串，如 "2020~2025", "2030", "全部"
            available_years: 可用的年份列表

        Returns:
            要修改的年份列表
        """
        year_spec = year_spec.strip()

        # 全部年份
        if year_spec in ['全部', '所有', '每年', '所有年份', '全部年份', 'all']:
            return available_years

        # 範圍：2020~2025, 2020-2025, 2020到2025, 從2020到2025
        range_patterns = [
            r'從?(\d{4})\s*[~\-到至]\s*(\d{4})',  # 2020~2025, 2020-2025, 2020到2025
            r'(\d{4})\s*年?\s*[~\-到至]\s*(\d{4})\s*年?',  # 2020年~2025年
        ]
        for pattern in range_patterns:
            match = re.search(pattern, year_spec)
            if match:
                start_year = int(match.group(1))
                end_year = int(match.group(2))
                return [y for y in available_years if start_year <= y <= end_year]

        # 開放範圍：2025之後, 2025以後, 從2025開始
        after_patterns = [
            r'(\d{4})\s*年?\s*[之以]後',
            r'從\s*(\d{4})\s*年?\s*開始',
            r'(\d{4})\s*年?\s*以後',
        ]
        for pattern in after_patterns:
            match = re.search(pattern, year_spec)
            if match:
                start_year = int(match.group(1))
                return [y for y in available_years if y >= start_year]

        # 開放範圍：2025之前, 2025以前
        before_patterns = [
            r'(\d{4})\s*年?\s*[之以]前',
        ]
        for pattern in before_patterns:
            match = re.search(pattern, year_spec)
            if match:
                end_year = int(match.group(1))
                return [y for y in available_years if y <= end_year]

        # 單一年份：2030, 2030年
        single_year_match = re.search(r'(\d{4})', year_spec)
        if single_year_match:
            year = int(single_year_match.group(1))
            if year in available_years:
                return [year]

        return []

    def edit_by_field_and_year(
        self,
        sheet_name: str,
        field_keyword: str,
        year_spec: str,
        new_value: Union[int, float] = None,
        section_type: str = None,
        year_value_map: dict = None
    ) -> dict:
        """
        根據欄位名稱和年份智慧修改儲存格

        Args:
            sheet_name: 工作表名稱，如 "滾算紀錄5"
            field_keyword: 欄位關鍵字，如 "設備費用"（支援模糊匹配）
            year_spec: 年份規格，如 "2020~2025", "2030", "全部", "multiple"（多年份模式）
            new_value: 新的數值，正負號完全依照使用者輸入。當 year_value_map 有值時可為 None
            section_type: 區域類型，"公版"=1-36行，"現金流量表"或"綜合損益表"=37-64行，None=全部
            year_value_map: 年份-數值映射，如 {2020: -40000, 2023: -20000}，支持非連續年份

        Returns:
            操作結果字典
        """
        try:
            # 強制要求 section_type：避免 LLM 猜測錯誤區域
            if section_type is None:
                return {
                    "success": False,
                    "need_clarification": True,
                    "message": (
                        "請指定要修改的區域：\n"
                        "- 「公版」（第1-36行）\n"
                        "- 「綜合損益表」（第37-64行）\n"
                        "- 「現金流量表」（第86-115行）"
                    )
                }

            wb = load_workbook(self.file_path)

            # 檢查 sheet 是否存在
            if sheet_name not in wb.sheetnames:
                # 嘗試模糊匹配 sheet 名稱
                matched_sheet = None
                for name in wb.sheetnames:
                    if sheet_name in name or name in sheet_name:
                        matched_sheet = name
                        break
                if not matched_sheet:
                    return {
                        "success": False,
                        "message": f"找不到工作表「{sheet_name}」",
                        "available_sheets": wb.sheetnames
                    }
                sheet_name = matched_sheet

            ws = wb[sheet_name]

            # 根據 section_type 決定搜索範圍和年份行
            if section_type == "公版":
                row_start, row_end = 1, 36
                year_row = 37  # 公版沒有年份行，使用綜合損益表的年份行
            elif section_type == "綜合損益表":
                row_start, row_end = 37, 64
                year_row = 37  # 綜合損益表的年份在第 37 行
            elif section_type == "現金流量表":
                row_start, row_end = 86, 115
                year_row = 87  # 現金流量表的年份在第 87 行
            else:
                # 預設搜索全部範圍
                row_start, row_end = 1, ws.max_row
                year_row = 37  # 預設年份行

            # 1. 掃描 B 列找出指定範圍內的欄位名稱
            field_candidates = []
            for row in range(row_start, min(row_end + 1, ws.max_row + 1)):
                cell_value = ws.cell(row=row, column=2).value  # B 列
                if cell_value and isinstance(cell_value, str):
                    field_candidates.append((row, cell_value))

            if not field_candidates:
                section_info = f"（{section_type}區域，第{row_start}-{row_end}行）" if section_type else ""
                return {
                    "success": False,
                    "message": f"在工作表「{sheet_name}」的 B 列{section_info}中找不到任何欄位名稱"
                }

            # 2. 模糊匹配欄位
            match_result = self._fuzzy_match(field_keyword, field_candidates)
            if not match_result:
                field_names = [f[1] for f in field_candidates]
                section_info = f"（{section_type}區域）" if section_type else ""
                return {
                    "success": False,
                    "message": f"在{section_info}找不到與「{field_keyword}」匹配的欄位",
                    "available_fields": field_names[:20]  # 最多顯示 20 個
                }

            field_row, matched_field, similarity = match_result

            # 判斷是否為支出
            is_outflow = 'outflow' in matched_field.lower()
            modified_cells = []

            # 3. 公版特殊處理：只有一格（D 列），不需要年份
            if section_type == "公版":
                # 公版數值在 D 列（第 4 列）
                value_col = 4

                # 直接使用使用者輸入的數值（含正負號）
                if new_value is not None:
                    final_value = new_value
                else:
                    return {
                        "success": False,
                        "message": "公版修改需要提供數值"
                    }

                cell_ref = f"{get_column_letter(value_col)}{field_row}"
                old_value = ws.cell(row=field_row, column=value_col).value
                ws.cell(row=field_row, column=value_col).value = final_value
                modified_cells.append({
                    "cell": cell_ref,
                    "year": None,
                    "old_value": old_value,
                    "new_value": final_value
                })

                # 備註寫入（同一 wb，不另外 open/save）
                self._write_remark_to_wb(wb, sheet_name, field_keyword, final_value)

                # 儲存一次，recalc 一次
                wb.save(self.file_path)
                _recalc(self.file_path)

                return {
                    "success": True,
                    "message": f"已修改「{sheet_name}」公版的「{matched_field}」，新值：{final_value}",
                    "sheet_name": sheet_name,
                    "matched_field": matched_field,
                    "field_keyword": field_keyword,
                    "similarity_score": round(similarity, 2),
                    "section_type": "公版",
                    "cells_modified": 1,
                    "new_value": final_value,
                    "is_outflow": is_outflow,
                    "details": modified_cells
                }

            # 4. 非公版：掃描年份行找出年份和對應的欄位
            year_columns = {}  # {year: column_number}
            for col in range(4, ws.max_column + 1):  # 從 D 列開始
                cell_value = ws.cell(row=year_row, column=col).value
                if cell_value is not None:
                    try:
                        year = int(cell_value)
                        if 1900 < year < 2100:  # 合理的年份範圍
                            year_columns[year] = col
                    except (ValueError, TypeError):
                        continue

            if not year_columns:
                return {
                    "success": False,
                    "message": f"在工作表「{sheet_name}」的第 {year_row} 行找不到年份資料"
                }

            available_years = sorted(year_columns.keys())

            # 5. 判斷使用哪種模式：year_value_map（多年份不同值）或傳統模式
            if year_value_map:
                # 多年份-不同值模式
                print(f"使用 year_value_map 模式: {year_value_map}")

                for year, value in year_value_map.items():
                    year_int = int(year) if isinstance(year, str) else year
                    if year_int not in year_columns:
                        print(f"警告: 年份 {year_int} 不在可用年份中，跳過")
                        continue

                    col = year_columns[year_int]

                    # 直接使用使用者輸入的數值（含正負號）
                    final_value = value

                    cell_ref = f"{get_column_letter(col)}{field_row}"
                    old_value = ws.cell(row=field_row, column=col).value
                    ws.cell(row=field_row, column=col).value = final_value
                    modified_cells.append({
                        "cell": cell_ref,
                        "year": year_int,
                        "old_value": old_value,
                        "new_value": final_value
                    })

                years_to_modify = [int(y) if isinstance(y, str) else y for y in year_value_map.keys()]
            else:
                # 傳統模式：單一值套用到多個年份
                years_to_modify = self._parse_year_spec(year_spec, available_years)

                if not years_to_modify:
                    return {
                        "success": False,
                        "message": f"無法解析年份規格「{year_spec}」或指定的年份不存在",
                        "available_years": available_years
                    }

                # 直接使用使用者輸入的數值（含正負號）
                final_value = new_value

                # 修改儲存格
                for year in years_to_modify:
                    col = year_columns[year]
                    cell_ref = f"{get_column_letter(col)}{field_row}"
                    old_value = ws.cell(row=field_row, column=col).value
                    ws.cell(row=field_row, column=col).value = final_value
                    modified_cells.append({
                        "cell": cell_ref,
                        "year": year,
                        "old_value": old_value,
                        "new_value": final_value
                    })

            # 6. 組織回應訊息
            if year_value_map:
                years_values_str = ", ".join([f"{y}年={v}" for y, v in sorted(year_value_map.items())])
                message = f"已修改「{sheet_name}」的「{matched_field}」，共 {len(modified_cells)} 個儲存格\n設定值：{years_values_str}"
                sample_value = list(year_value_map.values())[0]
                final_value_for_return = sample_value
            else:
                years_str = f"{min(years_to_modify)}~{max(years_to_modify)}" if len(years_to_modify) > 1 else str(years_to_modify[0])
                message = f"已修改「{sheet_name}」的「{matched_field}」，年份 {years_str}，共 {len(modified_cells)} 個儲存格，新值：{final_value}"
                final_value_for_return = final_value

            # 備註寫入（同一 wb，save 前寫入，避免多次 save 清除公式快取）
            self._write_remark_to_wb(wb, sheet_name, field_keyword, final_value_for_return)

            # 5. 儲存一次，recalc 一次
            wb.save(self.file_path)
            _recalc(self.file_path)

            return {
                "success": True,
                "message": message,
                "sheet_name": sheet_name,
                "matched_field": matched_field,
                "field_keyword": field_keyword,
                "similarity_score": round(similarity, 2),
                "years_modified": sorted(years_to_modify),
                "cells_modified": len(modified_cells),
                "new_value": final_value_for_return,
                "year_value_map": year_value_map,
                "is_outflow": is_outflow,
                "details": modified_cells
            }

        except Exception as e:
            import traceback
            return {
                "success": False,
                "message": f"修改失敗: {str(e)}",
                "traceback": traceback.format_exc()
            }

    # ========== 備註追加 ==========

    def _write_remark_to_wb(self, wb, sheet_name: str, field_keyword: str, new_value):
        """若修改的是 pN 分頁，在已開啟的 wb 中追加備註到「滾算紀錄單(總紀錄)」。
        不自行 open/save，由呼叫端統一存檔，避免多次 save 清除公式快取。"""
        import re
        from datetime import datetime

        m = re.match(r'^p(\d+)$', sheet_name, re.IGNORECASE)
        if not m:
            return

        record_number = int(m.group(1))

        try:
            if "滾算紀錄單(總紀錄)" not in wb.sheetnames:
                return

            summary = wb["滾算紀錄單(總紀錄)"]

            target_row = None
            for row in range(2, summary.max_row + 1):
                if summary.cell(row=row, column=1).value == record_number:
                    target_row = row
                    break

            if target_row is None:
                return

            now = datetime.now()
            note = f"{now.month}/{now.day}，使用者將{field_keyword}調整為{new_value}"

            current = summary.cell(row=target_row, column=10).value
            summary.cell(row=target_row, column=10).value = (
                f"{current}\n{note}" if current else note
            )
            print(f"[備註] 已寫入 滾算紀錄單(總紀錄) 第 {target_row} 列：{note}")

        except Exception as e:
            print(f"[備註] 寫入備註時發生錯誤: {str(e)}")

    # ========== 查詢功能 ==========

    def list_sheets(self) -> dict:
        """
        列出目前 Excel 檔案中所有工作表名稱

        Returns:
            包含工作表名稱清單的結果字典
        """
        try:
            wb = load_workbook(self.file_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return {
                "success": True,
                "message": f"共找到 {len(sheets)} 個工作表",
                "sheets": sheets
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"讀取工作表清單失敗: {str(e)}"
            }

    def _format_query_value(self, value, field_keyword: str):
        """格式化查詢值：IRR 類保留兩位小數並轉百分比，其他四捨五入到整數"""
        if value is None:
            return None
        if not isinstance(value, (int, float)):
            return value
        fk_lower = field_keyword.lower()
        is_irr = 'irr' in fk_lower
        v = float(value)
        if is_irr:
            # 若以小數儲存（如 0.0272）則轉為百分比（2.72）
            if abs(v) < 1:
                v = v * 100
            return round(v, 2)
        return round(v)

    def read_sheet_by_field(
        self,
        sheet_name: str,
        field_keyword: str,
        section_type: str = None,
        year: str = None
    ) -> dict:
        """
        全表動態掃描，根據欄位名稱查詢數值。
        - 掃描整個 B 欄找欄位標籤（模糊匹配），不限定固定行範圍
        - 往上搜尋最近的年份行，動態確定年份對應欄位
        - 若找不到年份行，視為單一值欄位，讀 C 欄（D 欄備用）
        - section_type 保留參數但忽略（向下相容）

        Args:
            sheet_name: 工作表名稱
            field_keyword: B 欄標頭關鍵字，支援模糊匹配
            section_type: 忽略（向下相容保留）
            year_spec: 年份規格（選填）。不填 = 回傳全部年份；
                       「2025」= 只回傳該年；「2020~2025」= 回傳範圍
        """
        try:
            wb = load_workbook(self.file_path, data_only=True)

            # 模糊匹配 sheet 名稱
            if sheet_name not in wb.sheetnames:
                matched_sheet = None
                for name in wb.sheetnames:
                    if sheet_name in name or name in sheet_name:
                        matched_sheet = name
                        break
                if not matched_sheet:
                    wb.close()
                    return {
                        "success": False,
                        "message": f"找不到工作表「{sheet_name}」",
                        "available_sheets": list(wb.sheetnames)
                    }
                sheet_name = matched_sheet

            ws = wb[sheet_name]

            # 掃描整個 B 欄建立候選清單
            field_candidates = []
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=2).value
                if cell_value and isinstance(cell_value, str) and cell_value.strip():
                    field_candidates.append((row, cell_value.strip()))

            if not field_candidates:
                wb.close()
                return {
                    "success": False,
                    "message": f"在工作表「{sheet_name}」中找不到任何欄位標籤"
                }

            # 模糊匹配欄位
            match_result = self._fuzzy_match(field_keyword, field_candidates)
            if not match_result:
                wb.close()
                return {
                    "success": False,
                    "message": f"找不到與「{field_keyword}」匹配的欄位",
                    "available_fields": [f[1] for f in field_candidates[:20]]
                }

            field_row, matched_field, similarity = match_result

            # 往上搜尋最近的年份行（含 3 個以上年份值的行）
            year_columns = {}
            for r in range(field_row - 1, max(0, field_row - 50), -1):
                cols = {}
                for col in range(3, min(ws.max_column + 1, 70)):
                    val = ws.cell(row=r, column=col).value
                    if val is not None:
                        try:
                            y = int(float(str(val)))
                            if 1900 < y < 2100:
                                cols[y] = col
                        except (ValueError, TypeError):
                            pass
                if len(cols) >= 3:
                    year_columns = cols
                    break

            if year_columns:
                # 有年份行：逐年讀取
                available_years = sorted(year_columns.keys())
                if year:
                    years_to_read = self._parse_year_spec(year, available_years)
                    if not years_to_read:
                        wb.close()
                        return {
                            "success": False,
                            "message": f"年份「{year}」不存在，可用年份：{available_years[0]}~{available_years[-1]}"
                        }
                else:
                    years_to_read = available_years

                result_values = {}
                for yr in years_to_read:
                    col = year_columns[yr]
                    raw = ws.cell(row=field_row, column=col).value
                    result_values[yr] = self._format_query_value(raw, matched_field)

                # 年份欄位全為 null → 視為單一值欄位，改讀 C 欄
                if all(v is None for v in result_values.values()):
                    value = ws.cell(row=field_row, column=3).value
                    if value is None:
                        value = ws.cell(row=field_row, column=4).value
                    wb.close()
                    return {
                        "success": True,
                        "sheet_name": sheet_name,
                        "matched_field": matched_field,
                        "value": self._format_query_value(value, matched_field)
                    }

                wb.close()
                return {
                    "success": True,
                    "sheet_name": sheet_name,
                    "matched_field": matched_field,
                    "years_queried": years_to_read,
                    "values": result_values
                }
            else:
                # 無年份行：單一值，先讀 C 欄，空則讀 D 欄
                value = ws.cell(row=field_row, column=3).value
                if value is None:
                    value = ws.cell(row=field_row, column=4).value
                wb.close()
                return {
                    "success": True,
                    "sheet_name": sheet_name,
                    "matched_field": matched_field,
                    "value": self._format_query_value(value, matched_field)
                }

        except Exception as e:
            print(f"[read_sheet_by_field 錯誤] {str(e)}")
            return {
                "success": False,
                "message": "查詢失敗，請確認工作表是否正確。"
            }

    def query_financial_data(
        self,
        sheet_name: str,
        field_keyword: str,
        year: str = None
    ) -> dict:
        """
        查詢工作表財務數據。全表動態掃描，無需指定區域。
        直接委派給 read_sheet_by_field。

        Args:
            sheet_name: 工作表名稱，如「p1」、「工作表1」
            field_keyword: 欄位關鍵字，如「IRR」、「現金流」、「保險費」、「稅後淨利」
            year: （選填）年份規格。不填 = 全部年份；「2025」= 單年；「2020~2025」= 範圍
        """
        return self.read_sheet_by_field(sheet_name, field_keyword, year=year)


# 定義工具的 schema（用於 function calling）
EXCEL_TOOLS_SCHEMA = [
    {
        "type": "function",
        "function": {
            "name": "write_excel_cell",
            "description": "在 Excel 文件中寫入或修改儲存格的值。可以寫入單個儲存格（如 A1）或範圍（如 A1:B5）",
            "parameters": {
                "type": "object",
                "properties": {
                    "cell": {
                        "type": "string",
                        "description": "儲存格位置，如 'A1' 或範圍 'A1:B5'"
                    },
                    "value": {
                        "type": ["string", "number"],
                        "description": "要寫入的值，可以是文字或數字"
                    }
                },
                "required": ["cell", "value"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "read_excel_cell",
            "description": "從 Excel 文件中讀取儲存格的值。可以讀取單個儲存格（如 A1）或範圍（如 A1:B5）",
            "parameters": {
                "type": "object",
                "properties": {
                    "cell": {
                        "type": "string",
                        "description": "儲存格位置，如 'A1' 或範圍 'A1:B5'"
                    }
                },
                "required": ["cell"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "delete_excel_cell",
            "description": "清空 Excel 文件中儲存格的內容。可以清空單個儲存格（如 A1）或範圍（如 A1:B5）",
            "parameters": {
                "type": "object",
                "properties": {
                    "cell": {
                        "type": "string",
                        "description": "儲存格位置，如 'A1' 或範圍 'A1:B5'"
                    }
                },
                "required": ["cell"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "edit_sheet_by_field",
            "description": "根據欄位名稱和年份智慧修改 Excel 工作表中的數值。支援模糊匹配欄位名稱。【區域限定】「公版」=1-36行，「綜合損益表」=37-64行，「現金流量表」=86-115行。【多年份模式】可使用 year_value_map 設定不同年份的不同數值。",
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名稱，如「滾算紀錄5」、「滾算紀錄_1」"
                    },
                    "field_keyword": {
                        "type": "string",
                        "description": "欄位關鍵字，如「設備費用」、「租金」、「保險費」等。支援模糊匹配"
                    },
                    "year_spec": {
                        "type": "string",
                        "description": "年份規格。支援：單一年份（「2030」）、範圍（「2020~2025」）、全部（「全部」）、多年份（「multiple」）。使用 year_value_map 時設為「multiple」"
                    },
                    "new_value": {
                        "type": "number",
                        "description": "要設定的新數值。正負號完全依照使用者輸入，使用者說 40000 就存 40000，說 -40000 就存 -40000。使用 year_value_map 時可省略"
                    },
                    "section_type": {
                        "type": "string",
                        "enum": ["公版", "綜合損益表", "現金流量表"],
                        "description": "【區域限定】「公版」→ 搜尋1-36行；「綜合損益表」→ 搜尋37-64行；「現金流量表」→ 搜尋86-115行。未提及則不填"
                    },
                    "year_value_map": {
                        "type": "object",
                        "description": "【多年份模式】年份與數值的映射，如 {\"2020\": -40000, \"2023\": -20000, \"2027\": -50000}。使用此參數時 year_spec 設為「multiple」，new_value 可省略"
                    }
                },
                "required": ["sheet_name", "field_keyword", "year_spec"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "list_excel_sheets",
            "description": "列出目前 Excel 檔案中所有工作表的名稱。當使用者詢問有哪些工作表、哪些滾算紀錄，或不確定工作表名稱時使用。",
            "parameters": {
                "type": "object",
                "properties": {},
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "read_sheet_by_field",
            "description": "根據欄位名稱查詢工作表中的數值。全表動態掃描，無需指定區域。支援模糊匹配。",
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名稱，可先用 list_excel_sheets 確認可用的名稱"
                    },
                    "field_keyword": {
                        "type": "string",
                        "description": "欄位關鍵字，對應 B 欄的標頭文字，如「專案法IRR」、「現金流」、「租金」。支援模糊匹配"
                    },
                    "year": {
                        "type": "string",
                        "description": "（選填）年份規格。不填 = 回傳全部年份；「2025」= 只回傳該年；「2020~2025」= 回傳範圍"
                    }
                },
                "required": ["sheet_name", "field_keyword"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "query_financial_data",
            "description": (
                "查詢工作表中的財務數據，自動判斷對應區域，無需指定 section_type。"
                "適用於查詢 IRR、現金流、稅後淨利、保險費、租金等任意欄位。"
                "支援所有工作表（工作表1、p1、p2 等）。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名稱，如「p1」、「p3」、「工作表1」"
                    },
                    "field_keyword": {
                        "type": "string",
                        "description": (
                            "欄位關鍵字，支援模糊匹配。例如：「IRR」、「專案法IRR」、"
                            "「現金流」、「稅後淨利」、「保險費」、「租金」、「運維費」"
                        )
                    },
                    "year": {
                        "type": "string",
                        "description": (
                            "（選填）年份規格。不填 = 回傳全部年份；"
                            "「2025」= 只回傳該年；「2020~2025」= 回傳範圍。"
                            "IRR 等單一值欄位不需填"
                        )
                    }
                },
                "required": ["sheet_name", "field_keyword"]
            }
        }
    }
]
