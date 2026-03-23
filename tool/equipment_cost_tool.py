# tool/equipment_cost_tool.py
import os
import shutil
import time
import datetime
from openpyxl import load_workbook, Workbook
from utils.recalc import recalc as _recalc
from typing import Dict, List, Optional, Any, Union
from tool.equipment_cost_services import (
    CostStructureService, 
    CashMode, 
    RatioMode, 
    ConditionalMode, 
    CustomizeMode
)
from tool.finance_tool import FinanceTool

class EquipmentCostTool:
    """設備成本滾算工具類"""

    def __init__(self):
        """初始化設備成本工具"""
        # 以此檔案位置為基準，往上一層找到專案根目錄，確保模板路徑正確
        _project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.excel_folder = os.path.join(_project_root, "Excel")
        self.excel_final_folder = os.path.join(_project_root, "Excel Generic Template")
        self.output_file_name = "excel公版.xlsx"

    def _find_excel_file(self, directory: str) -> Optional[str]:
        """在指定目錄中找到第一個 Excel 檔案"""
        if not os.path.exists(directory):
            return None

        for file in os.listdir(directory):
            if file.endswith(('.xlsx', '.xls')) and not file.startswith('~$'):
                return os.path.join(directory, file)
        return None

    def _safe_float_convert(self, value, default: float) -> float:
        """安全地將值轉換為浮點數"""
        try:
            if value is None:
                return default
            if isinstance(value, str) and (value.strip() == '' or '參數數值' in value):
                return default
            return float(value)
        except (ValueError, TypeError):
            return default
    
    def _safe_int_convert(self, value, default: int) -> int:
        """安全地將值轉換為整數"""
        try:
            if value is None:
                return default
            if isinstance(value, str) and (value.strip() == '' or '參數數值' in value):
                return default
            return int(float(value))
        except (ValueError, TypeError):
            return default

    def _read_excel_data(self, file_path: str) -> Dict[str, Any]:
        """讀取 Excel 檔案數據"""
        try:
            wb = load_workbook(file_path, data_only=True)
            sheet = wb.active

            # 根據提供的 Excel 結構讀取數據
            # 欄位對應：B欄為參數數值、C欄為額外數值、D欄為支提年限
            data = {
                "project_name": sheet["B2"].value,
                "plant_lifetime": self._safe_int_convert(sheet["B3"].value, 20),  # 電站壽命
                "capacity": self._safe_float_convert(sheet["B4"].value, 436.1),#建置量(kWp)
                "start_year": self._safe_int_convert(sheet["B5"].value, 2020),#起始年度
                "end_year": self._safe_int_convert(sheet["D5"].value, 2039),#結束年度
                "annual_generation": self._safe_float_convert(sheet["B10"].value, 0),#年度AC併網總發電量
                "annual_ac_generation": self._safe_float_convert(sheet["C10"].value, 0),  # 年度AC併網總發電量
                "electricity_revenue_years": self._safe_int_convert(sheet["D10"].value, 0),  # 電費收入年限
                "first_year_decline_rate": self._safe_float_convert(sheet["C11"].value, 0),  # 首年衰退率
                "fit_rate": self._safe_float_convert(sheet["C12"].value, 0),#次年衰退率
                "fit_price_c13": self._safe_float_convert(sheet["C13"].value, 0),  # 躉售費率(FIT)從C13讀取
                "equipment_cost": self._safe_int_convert(sheet["C16"].value, 0),  # 每kWp設備成本
                "equipment_amortization_years": self._safe_int_convert(sheet["D16"].value, 0),  # 設備費用(價金)支提年限
                "loan_interest": self._safe_float_convert(sheet["C26"].value, 0),#所得稅率
                "development_fee": self._safe_int_convert(sheet["C18"].value, 0),  # 開發費
                "rent_calculation_method": self._safe_float_convert(sheet["C19"].value, 0),  # 租金費用租金計算方式
                "rent_calculation_amortization_years": self._safe_int_convert(sheet["D19"].value, 0),  # 租金費用租金計算方式支提年限
                "rent_method1_total": self._safe_float_convert(sheet["C20"].value, 0),  # 租金費用【方式1參數】租金總額
                "rent_method1_amortization_years": self._safe_int_convert(sheet["D20"].value, 0),  # 租金費用【方式1參數】租金總額支提年限
                "rent_method2_ratio": self._safe_float_convert(sheet["C21"].value, 0),  # 租金費用【方式2參數】抽成比率
                "rent_method2_amortization_years": self._safe_int_convert(sheet["D21"].value, 0),  # 租金費用【方式2參數】抽成比率支提年限
                "loan_amount": self._safe_float_convert(sheet["C24"].value, 0),#【方式1參數】租金總額
                "maintenance_cost": self._safe_float_convert(sheet["C22"].value, 0),#運維費用每kw運維成本
                "maintenance_amortization_years": self._safe_int_convert(sheet["D22"].value, 0),  # 運維費用支提年限
                "insurance_cost": self._safe_float_convert(sheet["C23"].value, 0),#保險費用占整體設備費用比例
                "insurance_amortization_years": self._safe_int_convert(sheet["D23"].value, 0),  # 保險費用支提年限
                "recycle_cost": self._safe_float_convert(sheet["C24"].value, 0),  # 從C24讀取回收費用
                "recycle_amortization_years": self._safe_int_convert(sheet["D24"].value, 0),  # 模組回收費用支提年限
                "profit_rate": self._safe_float_convert(sheet["C17"].value, 0),  # 信邦利潤率
                "bank_interest_rate": self._safe_float_convert(sheet["C25"].value, 0),  # 利息費用銀行利率
                "interest_amortization_years": self._safe_int_convert(sheet["D25"].value, 0),  # 利息費用支提年限
                "tax_rate": self._safe_float_convert(sheet["C26"].value, 0),  # 所得稅率
                "tax_amortization_years": self._safe_int_convert(sheet["D26"].value, 0),  # 所得稅支提年限
                # C31~C34 數據讀取
                "c31_value": self._safe_float_convert(sheet["C31"].value, 0),  # 貸款成數占總設備費用比例
                "c32_value": self._safe_float_convert(sheet["C32"].value, 0),  # 貸款還款攤還期數
                "c33_value": self._safe_float_convert(sheet["C33"].value, 0),  # 現金股利股利比率
                "c34_value": self._safe_float_convert(sheet["C34"].value, 0)   # 年底減資攤還期數
            }

            wb.close()
            return data
            
        except Exception as e:
            raise Exception(f"讀取 Excel 檔案失敗: {str(e)}")

    def _calculate_irr_for_prices(self, excel_file: str, adjustment_record: List[int], profit_rate: float, development_fee: int, sheet_name: str = None) -> List[Dict[str, float]]:
        """為每個價金計算IRR（參考price_rolling_tool實作）"""
        try:
            # 每次都創建新的 FinanceTool 實例，避免多聊天室之間的狀態污染
            finance_tool = FinanceTool()

            # 設定Excel檔案
            finance_tool.set_excel_file(excel_file, sheet_name)

            print(f"[DEBUG] 使用Excel檔案: {excel_file}")
            if sheet_name:
                print(f"[DEBUG] 使用工作表: {sheet_name}")

            irr_results = []

            for price in adjustment_record:
                try:
                    # 使用財務工具計算IRR（參考price_rolling_tool的方式）
                    irr_result = finance_tool.calculate_scenario_irr(
                        equipment_cost=price,
                        profit_rate=profit_rate,
                        development_cost=development_fee,
                        sheet_name=sheet_name
                    )
                    irr_results.append(irr_result)
                except Exception as e:
                    print(f"計算 IRR 失敗 (價金 {price}): {str(e)}")
                    # 如果計算失敗，填入空值
                    irr_results.append({
                        "project_irr": None,
                        "cost_method_irr": None,
                        "equity_method_irr": None
                    })

            return irr_results

        except Exception as e:
            print(f"初始化財務工具失敗: {str(e)}")
            # 如果初始化失敗，為所有價金返回空值
            return [{
                "project_irr": None,
                "cost_method_irr": None,
                "equity_method_irr": None
            } for _ in adjustment_record]

    def _prepare_output_file(self, source_file_path: str) -> str:
        """直接返回輸入檔案路徑，在原檔案上操作"""
        # 確保 Excel Generic Template 資料夾存在（用於存放模板）
        if not os.path.exists(self.excel_final_folder):
            os.makedirs(self.excel_final_folder)

        print(f"將直接在輸入檔案上進行滾算記錄: {source_file_path}")
        return source_file_path

    def _get_next_record_number(self, wb) -> int:
        """取得下一個可用的滾算紀錄編號（統一編號系統）"""
        import re
        existing_numbers = []
        for sheet_name in wb.sheetnames:
            # 檢查「pN」格式，如 p1, p2, p3
            m = re.match(r'^p(\d+)$', sheet_name, re.IGNORECASE)
            if m:
                try:
                    existing_numbers.append(int(m.group(1)))
                except ValueError:
                    pass

        # 返回下一個可用編號
        if not existing_numbers:
            return 1
        else:
            return max(existing_numbers) + 1

    def _fill_copied_sheet_with_data(self, new_sheet, result: Dict[str, Any], profit_rate: float, capacity: float):
        """在複製的工作表中填入滾算結果數據，為後續從輸入公版取值做準備"""
        try:
            # 基本滾算資訊
            adjustment_record = result.get("adjustment_record", [])
            irr_results = result.get("irr_results", [])
            mode = result.get("mode", "")
            boundary = result.get("boundary", "")
            
            # 找一個不會影響原有格式的位置來放滾算結果數據
            # 在工作表的右側或下方空白區域填入數據，避免覆蓋原有內容
            
            # 先找到工作表的最大使用範圍
            max_row = new_sheet.max_row
            max_col = new_sheet.max_column
            
            # 在右側空白區域(例如從列Z開始)填入滾算摘要資訊
            summary_col = max_col + 2  # 留兩列間距
            
            new_sheet.cell(row=1, column=summary_col, value="滾算結果摘要")
            new_sheet.cell(row=2, column=summary_col, value=f"模式: {mode}")
            new_sheet.cell(row=3, column=summary_col, value=f"邊界值: {boundary}")  
            new_sheet.cell(row=4, column=summary_col, value=f"利潤率: {profit_rate}")
            new_sheet.cell(row=5, column=summary_col, value=f"建置容量: {capacity}")
            
            # 在下方填入價金和IRR序列
            if adjustment_record:
                new_sheet.cell(row=7, column=summary_col, value="滾算價金:")
                for i, price in enumerate(adjustment_record[:5]):  # 只顯示前5個
                    new_sheet.cell(row=8+i, column=summary_col, value=price)
            
            if irr_results:
                new_sheet.cell(row=7, column=summary_col+1, value="專案IRR:")
                for i, irr_data in enumerate(irr_results[:5]):  # 只顯示前5個
                    project_irr = irr_data.get("project_irr")
                    if project_irr:
                        new_sheet.cell(row=8+i, column=summary_col+1, value=f"{project_irr:.2f}%")
            
            print(f"已在 {new_sheet.title} 工作表右側空白區域填入滾算結果摘要")

        except Exception as e:
            print(f"填入複製工作表數據時發生錯誤: {str(e)}")


    def _auto_adjust_column_widths(self, sheet, min_width=8, max_width=50):
        """自動調整工作表的欄寬

        Args:
            sheet: 要調整的工作表
            min_width: 最小欄寬（默認8）
            max_width: 最大欄寬（默認50）
        """
        try:
            from openpyxl.utils import get_column_letter

            for col_idx in range(1, sheet.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)

                for row_idx in range(1, min(sheet.max_row + 1, 120)):  # 只檢查前120行
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        # 計算單元格內容的長度（考慮中文字符寬度）
                        cell_value = str(cell.value)
                        # 中文字符算2個字符寬度
                        length = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                        max_length = max(max_length, length)

                # 設置欄寬，添加一些緩衝
                adjusted_width = max(min_width, min(max_length + 2, max_width))
                sheet.column_dimensions[column_letter].width = adjusted_width

            print(f"已自動調整 {sheet.title} 工作表的欄寬")
        except Exception as e:
            print(f"自動調整欄寬時發生錯誤: {str(e)}")

    def _fill_record_sheet_with_data(self, new_sheet, result, rolling_idx, record_number):
        """為滾算紀錄工作表填入數據（統一方法，用於所有滾算紀錄）

        Args:
            new_sheet: 要填入數據的工作表
            result: 滾算結果數據
            rolling_idx: 滾算索引（0=初始價金, 1=第1次滾算...）
            record_number: 滾算紀錄編號（對應滾算紀錄單的編號）
        """
        try:
            adjustment_record = result.get("adjustment_record", [])
            profit_record = result.get("profit_record", [])
            irr_results = result.get("irr_results", [])
            mode = result.get("mode", "")
            step = result.get("step", "")
            boundary = result.get("boundary", "")
            original_data = result.get("original_data", {})
            user_development_fee = result.get("development_fee", original_data.get("development_fee", 0))

            # === 1. 從輸入公版填入基礎資料到對應位置 ===
            project_name = original_data.get("project_name", "")
            if project_name:
                new_sheet.cell(row=2, column=3, value=project_name)  # C2

            new_sheet.cell(row=3, column=3, value=original_data.get("plant_lifetime", 20))  # C3 電站壽命
            new_sheet.cell(row=4, column=3, value=original_data.get("capacity", 436.1))  # C4 裝置容量
            new_sheet.cell(row=5, column=3, value=original_data.get("start_year", 2020))  # C5 起始年度
            new_sheet.cell(row=5, column=5, value=original_data.get("end_year", 2039))  # E5 結束年度

            new_sheet.cell(row=10, column=4, value=original_data.get("annual_ac_generation", 0))  # D10 年度AC併網總發電量
            new_sheet.cell(row=10, column=5, value=original_data.get("electricity_revenue_years", 0))  # E10 電費收入年限
            new_sheet.cell(row=11, column=4, value=original_data.get("first_year_decline_rate", 0))  # D11 首年衰退率
            new_sheet.cell(row=12, column=4, value=original_data.get("fit_rate", 0))  # D12 次年衰退率
            new_sheet.cell(row=13, column=4, value=original_data.get("fit_price_c13", 0))  # D13 躉售費率

            # C17~C26 填到 D17~D26
            new_sheet.cell(row=18, column=4, value=user_development_fee)  # D18 開發費
            new_sheet.cell(row=19, column=4, value=original_data.get("rent_calculation_method", 0))  # D19
            new_sheet.cell(row=19, column=5, value=original_data.get("rent_calculation_amortization_years", 0))  # E19
            new_sheet.cell(row=20, column=4, value=original_data.get("rent_method1_total", 0))  # D20
            new_sheet.cell(row=20, column=5, value=original_data.get("rent_method1_amortization_years", 0))  # E20
            new_sheet.cell(row=21, column=4, value=original_data.get("rent_method2_ratio", 0))  # D21
            new_sheet.cell(row=21, column=5, value=original_data.get("rent_method2_amortization_years", 0))  # E21
            new_sheet.cell(row=22, column=4, value=original_data.get("maintenance_cost", 0))  # D22 運維費用
            new_sheet.cell(row=22, column=5, value=original_data.get("maintenance_amortization_years", 0))  # E22
            new_sheet.cell(row=23, column=4, value=original_data.get("insurance_cost", 0))  # D23 保險費用
            new_sheet.cell(row=23, column=5, value=original_data.get("insurance_amortization_years", 0))  # E23
            new_sheet.cell(row=24, column=4, value=original_data.get("loan_amount", 0))  # D24
            new_sheet.cell(row=24, column=5, value=original_data.get("recycle_amortization_years", 0))  # E24
            new_sheet.cell(row=25, column=4, value=original_data.get("bank_interest_rate", 0))  # D25 銀行利率
            new_sheet.cell(row=25, column=5, value=original_data.get("interest_amortization_years", 0))  # E25
            new_sheet.cell(row=26, column=4, value=original_data.get("tax_rate", 0))  # D26 所得稅率
            new_sheet.cell(row=26, column=5, value=original_data.get("tax_amortization_years", 0))  # E26

            # C31~C34 填到 D31~D34
            new_sheet.cell(row=31, column=4, value=original_data.get("c31_value", 0))  # D31
            new_sheet.cell(row=32, column=4, value=original_data.get("c32_value", 0))  # D32
            new_sheet.cell(row=33, column=4, value=original_data.get("c33_value", 0))  # D33
            new_sheet.cell(row=34, column=4, value=original_data.get("c34_value", 0))  # D34

            # === 2. 填入該次滾算的數據 ===
            if rolling_idx < len(adjustment_record):
                current_price = adjustment_record[rolling_idx]
                irr_data = irr_results[rolling_idx] if rolling_idx < len(irr_results) else {"project_irr": None, "equity_method_irr": None}

                # 獲取利潤率
                current_profit_rate = original_data.get("profit_rate", 0.2)
                if hasattr(self, 'current_profit_rate') and self.current_profit_rate is not None:
                    current_profit_rate = self.current_profit_rate

                # 填入設備成本和利潤率
                new_sheet.cell(row=16, column=4, value=current_price)  # D16 設備成本
                new_sheet.cell(row=16, column=5, value=original_data.get("equipment_amortization_years", 0))  # E16 支提年限
                new_sheet.cell(row=17, column=4, value=current_profit_rate)  # D17 信邦利潤率

                # === 3. 在右側空白區域填入滾算摘要 ===
                summary_col = 13  # 列M
                new_sheet.cell(row=4, column=summary_col, value=f"情境編號: {record_number}")

                if rolling_idx == 0:
                    new_sheet.cell(row=5, column=summary_col, value="初始價金摘要")
                else:
                    new_sheet.cell(row=5, column=summary_col, value=f"第{rolling_idx}次滾算摘要")

                new_sheet.cell(row=6, column=summary_col, value=f"價金: {current_price}")
                new_sheet.cell(row=7, column=summary_col, value=f"專案IRR: {irr_data.get('project_irr', 'N/A')}")
                new_sheet.cell(row=8, column=summary_col, value=f"權益法IRR: {irr_data.get('equity_method_irr', 'N/A')}")
                new_sheet.cell(row=9, column=summary_col, value=f"利潤率: {current_profit_rate*100:.2f}%")

                # 生成備註說明
                mode_desc = {
                    "cash": f"現金模式，每次減少 {step} 元",
                    "ratio": f"比率模式，每次減少 {step*100}%",
                    "conditional": "條件模式，依價金範圍調整步幅",
                    "customize": "自訂模式，自動或手動配置步幅"
                }.get(mode, f"{mode}模式")

                if rolling_idx == 0:
                    remark = f"初始價金 - {mode_desc}"
                else:
                    reduction = adjustment_record[rolling_idx-1] - current_price
                    remark = f"第{rolling_idx}次滾算，減少 {reduction} 元 - {mode_desc}"

                new_sheet.cell(row=10, column=summary_col, value=remark)

                # === 4. 填入金流明細資料（以負值形式）===
                cash_flow_details = irr_data.get("cash_flow_details", {})
                if cash_flow_details:
                    # D~W 對應 column 4~23 (共20欄)
                    start_col = 4  # D欄
                    max_cols = 20  # D到W共20欄
                    total_years = cash_flow_details.get("total_years", 20)

                    # 年份: D37~W37 (row 37) - 根據起始年和結束年填入
                    start_year = original_data.get("start_year", 2020)
                    for i in range(min(total_years, max_cols)):
                        new_sheet.cell(row=37, column=start_col + i, value=start_year + i)

                    # 租金費用: D45~W45 (row 45)
                    rent_list = cash_flow_details.get("rent_list", [])
                    for i, value in enumerate(rent_list[:max_cols]):
                        new_sheet.cell(row=45, column=start_col + i, value=-abs(value) if value else 0)

                    # 運維費用: D48~W48 (row 48)
                    maintenance_list = cash_flow_details.get("maintenance_list", [])
                    for i, value in enumerate(maintenance_list[:max_cols]):
                        new_sheet.cell(row=48, column=start_col + i, value=-abs(value) if value else 0)

                    # 保險費用: D49~W49 (row 49)
                    insurance_list = cash_flow_details.get("insurance_list", [])
                    for i, value in enumerate(insurance_list[:max_cols]):
                        new_sheet.cell(row=49, column=start_col + i, value=-abs(value) if value else 0)

                    # 回收費用: D50~W50 (row 50)
                    recycle_list = cash_flow_details.get("recycle_list", [])
                    for i, value in enumerate(recycle_list[:max_cols]):
                        new_sheet.cell(row=50, column=start_col + i, value=-abs(value) if value else 0)

                    # 貸款還款: E93~W93 (row 93, 從E欄開始)
                    pay_back_list = cash_flow_details.get("pay_back_list", [])
                    pay_back_start_col = 5  # E欄
                    for i, value in enumerate(pay_back_list[:max_cols - 1]):  # 少一欄因為從E開始
                        new_sheet.cell(row=93, column=pay_back_start_col + i, value=-abs(value) if value else 0)

                    # 年底減資: D96~W96 (row 96)
                    dividend_list = cash_flow_details.get("dividend_list", [])
                    for i, value in enumerate(dividend_list[:max_cols]):
                        new_sheet.cell(row=96, column=start_col + i, value=-abs(value) if value else 0)

                    print(f"已在 {new_sheet.title} 工作表填入金流明細資料")

                print(f"已在 {new_sheet.title} 工作表填入滾算紀錄 {record_number} 的數據")

        except Exception as e:
            print(f"填入滾算紀錄工作表數據時發生錯誤: {str(e)}")

    def _fill_copied_sheet_with_rolling_data(self, new_sheet, result, rolling_idx):
        """【已棄用】為特定的滾算次數在複製工作表中填入數據，並從輸入公版填入基礎資料"""
        try:
            adjustment_record = result.get("adjustment_record", [])
            profit_record = result.get("profit_record", [])
            irr_results = result.get("irr_results", [])
            mode = result.get("mode", "")
            step = result.get("step", "")
            boundary = result.get("boundary", "")
            original_data = result.get("original_data", {})
            user_development_fee = result.get("development_fee", original_data.get("development_fee", 0))
            
            # === 1. 從輸入公版填入基礎資料到對應位置 ===
            # 輸入公版的B2(項目名稱)填到C2
            project_name = original_data.get("project_name", "")
            if project_name:
                new_sheet.cell(row=2, column=3, value=project_name)  # C2
            
            # 輸入公版的B3(電站壽命)填到C3
            plant_lifetime = original_data.get("plant_lifetime", 20)
            new_sheet.cell(row=3, column=3, value=plant_lifetime)  # C3
            
            # 輸入公版的B4(裝置容量)填到C4 
            capacity = original_data.get("capacity", 436.1)
            new_sheet.cell(row=4, column=3, value=capacity)  # C4
            
            # 輸入公版的B5(起始年度)填到C5
            start_year = original_data.get("start_year", 2020)
            new_sheet.cell(row=5, column=3, value=start_year)  # C5
            
            # 輸入公版的D5(結束年度)填到E5
            end_year = original_data.get("end_year", 2039)
            new_sheet.cell(row=5, column=5, value=end_year)  # E5
            
            # 輸入公版的C10(年度AC併網總發電量)填到D10
            annual_ac_generation = original_data.get("annual_ac_generation", 547002)
            new_sheet.cell(row=10, column=4, value=annual_ac_generation)  # D10
            
            # 輸入公版的D10(電費收入年限)填到E10
            electricity_revenue_years = original_data.get("electricity_revenue_years", 0)
            new_sheet.cell(row=10, column=5, value=electricity_revenue_years)  # E10
            
            # 輸入公版的C11(首年衰退率)填到D11
            first_year_decline_rate = original_data.get("first_year_decline_rate", 0)
            new_sheet.cell(row=11, column=4, value=first_year_decline_rate)  # D11
            
            # 輸入公版的C12(次年衰退率)填到D12
            second_year_decline_rate = original_data.get("fit_rate", 0)
            new_sheet.cell(row=12, column=4, value=second_year_decline_rate)  # D12
            
            # 輸入公版的C13(躉售費率FIT)填到D13
            fit_price_c13 = original_data.get("fit_price_c13", 3.7152)
            new_sheet.cell(row=13, column=4, value=fit_price_c13)  # D13
            
            # 輸入公版的C17~C26填到D17~D26 (使用原有字段名稱)
            new_sheet.cell(row=17, column=4, value=original_data.get("profit_rate", 0))  # D17信邦利潤率
            new_sheet.cell(row=18, column=4, value=user_development_fee)  # D18開發費（使用用戶輸入值）
            new_sheet.cell(row=19, column=4, value=original_data.get("rent_calculation_method", 0))  # D19租金費用租金計算方式
            new_sheet.cell(row=19, column=5, value=original_data.get("rent_calculation_amortization_years", 0))  # E19租金費用租金計算方式支提年限
            new_sheet.cell(row=20, column=4, value=original_data.get("rent_method1_total", 0))  # D20租金費用【方式1參數】租金總額
            new_sheet.cell(row=20, column=5, value=original_data.get("rent_method1_amortization_years", 0))  # E20租金費用【方式1參數】租金總額支提年限
            new_sheet.cell(row=21, column=4, value=original_data.get("rent_method2_ratio", 0))  # D21租金費用【方式2參數】拽成比率
            new_sheet.cell(row=21, column=5, value=original_data.get("rent_method2_amortization_years", 0))  # E21租金費用【方式2參數】拽成比率支提年限
            new_sheet.cell(row=22, column=4, value=original_data.get("maintenance_cost", 0))  # D22運維費用
            new_sheet.cell(row=22, column=5, value=original_data.get("maintenance_amortization_years", 0))  # E22運維費用支提年限
            new_sheet.cell(row=23, column=4, value=original_data.get("insurance_cost", 0))  # D23保險費用
            new_sheet.cell(row=23, column=5, value=original_data.get("insurance_amortization_years", 0))  # E23保險費用支提年限
            new_sheet.cell(row=24, column=4, value=original_data.get("loan_amount", 0))  # D24租金總額
            new_sheet.cell(row=24, column=5, value=original_data.get("recycle_amortization_years", 0))  # E24模組回收費用支提年限
            new_sheet.cell(row=25, column=4, value=original_data.get("bank_interest_rate", 0))  # D25利息費用銀行利率
            new_sheet.cell(row=25, column=5, value=original_data.get("interest_amortization_years", 0))  # E25利息費用支提年限
            new_sheet.cell(row=26, column=4, value=original_data.get("tax_rate", 0))  # D26所得稅率
            new_sheet.cell(row=26, column=5, value=original_data.get("tax_amortization_years", 0))  # E26所得稅支提年限
            
            # 新增：輸入公版C31~C34填到D31~D34
            new_sheet.cell(row=31, column=4, value=original_data.get("c31_value", 0))  # D31: 貸款成數占總設備費用比例
            new_sheet.cell(row=32, column=4, value=original_data.get("c32_value", 0))  # D32: 貸款還款攤還期數
            new_sheet.cell(row=33, column=4, value=original_data.get("c33_value", 0))  # D33: 現金股利股利比率
            new_sheet.cell(row=34, column=4, value=original_data.get("c34_value", 0))  # D34: 年底減資攤還期數
            
            # 獲取特定滾算次數的數據
            if rolling_idx < len(adjustment_record):
                current_price = adjustment_record[rolling_idx]
                current_profit = profit_record[rolling_idx] if rolling_idx < len(profit_record) else 0
                irr_data = irr_results[rolling_idx] if rolling_idx < len(irr_results) else {"project_irr": None, "equity_method_irr": None}
                
                # 獲取利潤率和容量（優先使用用戶輸入值）
                current_profit_rate = original_data.get("profit_rate", 0.2)
                if hasattr(self, 'current_profit_rate') and self.current_profit_rate is not None:
                    current_profit_rate = self.current_profit_rate
                
                # === 2. 填入滾算計算結果到設備成本位置 ===
                # 將當前滾算的設備成本填入對應位置（D16）
                new_sheet.cell(row=16, column=4, value=current_price)  # D16設備成本
                
                # 輸入公版的D16(設備費用攤提年限)填到E16
                equipment_amortization_years = original_data.get("equipment_amortization_years", 20)
                new_sheet.cell(row=16, column=5, value=equipment_amortization_years)  # E16
                
                # 輸入公版的D16(設備費用支提年限)填到E16
                equipment_amortization_years = original_data.get("equipment_amortization_years", 0)
                new_sheet.cell(row=16, column=5, value=equipment_amortization_years)  # E16
                
                # === 3. 填入利潤率（優先使用用戶輸入值） ===
                # 填入到D17信邦利潤率位置
                new_sheet.cell(row=17, column=4, value=current_profit_rate)  # D17信邦利潤率
                
                # === 4. 在右側空白區域填入滾算摘要（保持原有功能） ===
                summary_col = 13  # 列M
                new_sheet.cell(row=5, column=summary_col, value=f"第{rolling_idx}次滾算摘要")
                new_sheet.cell(row=6, column=summary_col, value=f"價金: {current_price}")
                new_sheet.cell(row=7, column=summary_col, value=f"專案IRR: {irr_data.get('project_irr', 'N/A')}")
                new_sheet.cell(row=8, column=summary_col, value=f"權益法IRR: {irr_data.get('equity_method_irr', 'N/A')}")
                new_sheet.cell(row=9, column=summary_col, value=f"利潤率: {current_profit_rate*100:.2f}%")
                
                # 生成備註說明
                mode_desc = {
                    "cash": f"現金模式，每次減少 {step} 元",
                    "ratio": f"比率模式，每次減少 {step*100}%",
                    "conditional": "條件模式，依價金範圍調整步幅",
                    "customize": "自訂模式，自動或手動配置步幅"
                }.get(mode, f"{mode}模式")
                
                if rolling_idx == 0:
                    remark = f"初始價金 - {mode_desc}"
                else:
                    reduction = adjustment_record[rolling_idx-1] - current_price
                    remark = f"第{rolling_idx}次滾算，減少 {reduction} 元 - {mode_desc}"
                
                new_sheet.cell(row=10, column=summary_col, value=remark)
                
                print(f"已在 {new_sheet.title} 工作表填入輸入公版基礎資料和第{rolling_idx}次滾算結果")
            
        except Exception as e:
            print(f"填入複製工作表滾算數據時發生錯誤: {str(e)}")

    def _fill_original_sheet_with_initial_data(self, wb, result):
        """【已棄用】為滾算紀錄工作表填入初始價金的數據"""
        try:
            # 獲取下一個滾算紀錄編號（使用統一的編號方法）
            base_record_number = self._get_next_record_number(wb)
            sheet_name = f"p{base_record_number}"

            # 創建或獲取工作表
            if sheet_name not in wb.sheetnames:
                # 嘗試從模板複製
                template_copied = False

                # 優先從輸入檔案中已有的「滾算紀錄1」複製
                if "滾算紀錄1" in wb.sheetnames:
                    sheet = wb.copy_worksheet(wb["滾算紀錄1"])
                    sheet.title = sheet_name
                    print(f"已從輸入檔案的「滾算紀錄1」複製為「{sheet_name}」")
                    template_copied = True
                else:
                    # 從 Excel Generic Template/excel公版.xlsx 複製模板
                    template_file = os.path.join(self.excel_final_folder, "excel公版.xlsx")
                    if os.path.exists(template_file):
                        try:
                            template_wb = load_workbook(template_file)
                            if "滾算紀錄1" in template_wb.sheetnames:
                                # 複製工作表內容
                                template_sheet = template_wb["滾算紀錄1"]
                                sheet = wb.create_sheet(sheet_name)

                                # 複製所有單元格內容和格式
                                for row in template_sheet.iter_rows():
                                    for cell in row:
                                        new_cell = sheet[cell.coordinate]
                                        new_cell.value = cell.value
                                        if cell.has_style:
                                            new_cell.font = cell.font.copy()
                                            new_cell.border = cell.border.copy()
                                            new_cell.fill = cell.fill.copy()
                                            new_cell.number_format = cell.number_format
                                            new_cell.protection = cell.protection.copy()
                                            new_cell.alignment = cell.alignment.copy()

                                template_wb.close()
                                print(f"已從模板檔案複製為「{sheet_name}」")
                                template_copied = True
                            else:
                                template_wb.close()
                        except Exception as e:
                            print(f"從模板檔案複製失敗: {str(e)}")

                # 如果沒有複製成功，創建空白工作表
                if not template_copied:
                    sheet = wb.create_sheet(sheet_name)
                    print(f"已創建新工作表「{sheet_name}」（無模板）")
            else:
                sheet = wb[sheet_name]
                print(f"工作表「{sheet_name}」已存在，將覆蓋數據")
            adjustment_record = result.get("adjustment_record", [])
            irr_results = result.get("irr_results", [])
            original_data = result.get("original_data", {})
            # 獲取用戶輸入的開發費（優先使用用戶輸入值）
            user_development_fee = result.get("development_fee", original_data.get("development_fee", 0))
            
            # === 1. 從輸入公版填入基礎資料到對應位置 ===
            # 輸入公版的B2(項目名稱)填到C2
            project_name = original_data.get("project_name", "")
            if project_name:
                sheet.cell(row=2, column=3, value=project_name)  # C2
            
            # 輸入公版的B3(電站壽命)填到C3
            plant_lifetime = original_data.get("plant_lifetime", 20)
            sheet.cell(row=3, column=3, value=plant_lifetime)  # C3
            
            # 輸入公版的B4(裝置容量)填到C4 
            capacity = original_data.get("capacity", 436.1)
            sheet.cell(row=4, column=3, value=capacity)  # C4
            
            # 輸入公版的B5(起始年度)填到C5
            start_year = original_data.get("start_year", 2020)
            sheet.cell(row=5, column=3, value=start_year)  # C5
            
            # 輸入公版的D5(結束年度)填到E5
            end_year = original_data.get("end_year", 2039)
            sheet.cell(row=5, column=5, value=end_year)  # E5
            
            # 輸入公版的C10(年度AC併網總發電量)填到D10
            annual_ac_generation = original_data.get("annual_ac_generation", 547002)
            sheet.cell(row=10, column=4, value=annual_ac_generation)  # D10
            
            # 輸入公版的D10(電費收入年限)填到E10
            electricity_revenue_years = original_data.get("electricity_revenue_years", 0)
            sheet.cell(row=10, column=5, value=electricity_revenue_years)  # E10
            
            # 輸入公版的C11(首年衰退率)填到D11
            first_year_decline_rate = original_data.get("first_year_decline_rate", 0)
            sheet.cell(row=11, column=4, value=first_year_decline_rate)  # D11
            
            # 輸入公版的C12(次年衰退率)填到D12
            second_year_decline_rate = original_data.get("fit_rate", 0)
            sheet.cell(row=12, column=4, value=second_year_decline_rate)  # D12
            
            # 輸入公版的C13(躉售費率FIT)填到D13
            fit_price_c13 = original_data.get("fit_price_c13", 3.7152)
            sheet.cell(row=13, column=4, value=fit_price_c13)  # D13
            
            # 輸入公版的C17~C26填到D17~D26 (使用原有字段名稱)
            sheet.cell(row=17, column=4, value=original_data.get("profit_rate", 0))  # D17信邦利潤率
            sheet.cell(row=18, column=4, value=user_development_fee)  # D18開發費（使用用戶輸入值）
            sheet.cell(row=19, column=4, value=original_data.get("rent_calculation_method", 0))  # D19租金費用租金計算方式
            sheet.cell(row=19, column=5, value=original_data.get("rent_calculation_amortization_years", 0))  # E19租金費用租金計算方式支提年限
            sheet.cell(row=20, column=4, value=original_data.get("rent_method1_total", 0))  # D20租金費用【方式1參數】租金總額
            sheet.cell(row=20, column=5, value=original_data.get("rent_method1_amortization_years", 0))  # E20租金費用【方式1參數】租金總額支提年限
            sheet.cell(row=21, column=4, value=original_data.get("rent_method2_ratio", 0))  # D21租金費用【方式2參數】拽成比率
            sheet.cell(row=21, column=5, value=original_data.get("rent_method2_amortization_years", 0))  # E21租金費用【方式2參數】拽成比率支提年限
            sheet.cell(row=22, column=4, value=original_data.get("maintenance_cost", 0))  # D22運維費用
            sheet.cell(row=22, column=5, value=original_data.get("maintenance_amortization_years", 0))  # E22運維費用支提年限
            sheet.cell(row=23, column=4, value=original_data.get("insurance_cost", 0))  # D23保險費用
            sheet.cell(row=23, column=5, value=original_data.get("insurance_amortization_years", 0))  # E23保險費用支提年限
            sheet.cell(row=24, column=4, value=original_data.get("loan_amount", 0))  # D24租金總額
            sheet.cell(row=24, column=5, value=original_data.get("recycle_amortization_years", 0))  # E24模組回收費用支提年限
            sheet.cell(row=25, column=4, value=original_data.get("bank_interest_rate", 0))  # D25利息費用銀行利率
            sheet.cell(row=25, column=5, value=original_data.get("interest_amortization_years", 0))  # E25利息費用支提年限
            sheet.cell(row=26, column=4, value=original_data.get("tax_rate", 0))  # D26所得稅率
            sheet.cell(row=26, column=5, value=original_data.get("tax_amortization_years", 0))  # E26所得稅支提年限
            
            # 輸入公版的C10(年度AC併網總發電量)填到D10
            annual_ac_generation = original_data.get("annual_ac_generation", 547002)
            sheet.cell(row=10, column=4, value=annual_ac_generation)  # D10
            
            # 新增：輸入公版C31~C34填到D31~D34
            sheet.cell(row=31, column=4, value=original_data.get("c31_value", 0))  # D31: 貸款成數占總設備費用比例
            sheet.cell(row=32, column=4, value=original_data.get("c32_value", 0))  # D32: 貸款還款攤還期數
            sheet.cell(row=33, column=4, value=original_data.get("c33_value", 0))  # D33: 現金股利股利比率
            sheet.cell(row=34, column=4, value=original_data.get("c34_value", 0))  # D34: 年底減資攤還期數
            
            # === 2. 填入初始價金數據（第0次，即初始價金） ===
            if adjustment_record and irr_results:
                initial_price = adjustment_record[0]
                initial_irr = irr_results[0] if len(irr_results) > 0 else {"project_irr": None, "equity_method_irr": None}
                
                # 獲取利潤率（優先使用用戶輸入值）
                current_profit_rate = original_data.get("profit_rate", 0.2)
                if hasattr(self, 'current_profit_rate') and self.current_profit_rate is not None:
                    current_profit_rate = self.current_profit_rate
                
                # === 3. 填入初始設備成本到D16 ===
                sheet.cell(row=16, column=4, value=initial_price)  # D16設備成本
                
                # 輸入公版的D16(設備費用攤提年限)填到E16
                equipment_amortization_years = original_data.get("equipment_amortization_years", 20)
                sheet.cell(row=16, column=5, value=equipment_amortization_years)  # E16
                
                # 輸入公版的D16(設備費用支提年限)填到E16
                equipment_amortization_years = original_data.get("equipment_amortization_years", 0)
                sheet.cell(row=16, column=5, value=equipment_amortization_years)  # E16
                
                # === 4. 填入利潤率到D17 ===
                sheet.cell(row=17, column=4, value=current_profit_rate)  # D17信邦利潤率
                
                # === 5. 在右側空白區域填入初始價金摘要 ===
                summary_col = 13  # 列M
                sheet.cell(row=5, column=summary_col, value="初始價金摘要")
                sheet.cell(row=6, column=summary_col, value=f"初始價金: {initial_price}")
                sheet.cell(row=7, column=summary_col, value=f"專案IRR: {initial_irr.get('project_irr', 'N/A')}")
                sheet.cell(row=8, column=summary_col, value=f"權益法IRR: {initial_irr.get('equity_method_irr', 'N/A')}")
                sheet.cell(row=9, column=summary_col, value=f"利潤率: {current_profit_rate*100:.2f}%")
                sheet.cell(row=10, column=summary_col, value="基準價金 - 滾算起始點")
                
                print(f"已在 滾算紀錄1 工作表填入輸入公版基礎資料和初始價金數據")
        
        except Exception as e:
            print(f"填入滾算紀錄1工作表數據時發生錯誤: {str(e)}")

    def _write_results_to_excel(self, output_path: str, results: List[Dict[str, Any]]):
        """將結果寫入滾算紀錄單(總紀錄)工作表，並複製滾算紀錄工作表"""
        max_retries = 3
        retry_delay = 1

        for attempt in range(max_retries):
            try:
                # 使用 keep_links=True 保留外部連結，不使用 data_only 以保留公式
                wb = load_workbook(output_path, keep_links=True)

                # === 1. 寫入滾算紀錄單(總紀錄)工作表 ===
                EXPECTED_HEADERS = ["編號", "價金每KW", "專案法IRR", "成本法IRR", "權益法IRR", "信邦利潤率", "信邦利潤", "每kw信邦利潤", "開發費", "備註"]

                if "滾算紀錄單(總紀錄)" in wb.sheetnames:
                    sheet = wb["滾算紀錄單(總紀錄)"]
                else:
                    # 如果沒有滾算紀錄單(總紀錄)，創建一個
                    sheet = wb.create_sheet("滾算紀錄單(總紀錄)")
                    for col, header in enumerate(EXPECTED_HEADERS, 1):
                        sheet.cell(row=1, column=col, value=header)

                # 檢查是否需要添加「編號」欄位（針對已存在的工作表）
                if sheet.cell(row=1, column=1).value != "編號":
                    sheet.insert_cols(1)
                    sheet.cell(row=1, column=1, value="編號")

                # 檢查是否需要插入「成本法IRR」欄位（舊工作表 D 欄是「權益法IRR」）
                if sheet.cell(row=1, column=4).value == "權益法IRR":
                    sheet.insert_cols(4)
                    sheet.cell(row=1, column=4, value="成本法IRR")

                # 找到下一個空行開始寫入（檢查編號欄和價金欄）
                start_row = 2
                while sheet.cell(row=start_row, column=1).value is not None or sheet.cell(row=start_row, column=2).value is not None:
                    start_row += 1

                # 寫入滾算結果，並為每次滾算創建對應的工作表
                current_row = start_row
                for result_index, result in enumerate(results):
                    adjustment_record = result.get("adjustment_record", [])
                    profit_record = result.get("profit_record", [])
                    irr_results = result.get("irr_results", [])
                    mode = result.get("mode", "")
                    step = result.get("step", "")
                    boundary = result.get("boundary", "")
                    development_fee = result.get("development_fee", 0)

                    # 生成備註說明
                    mode_desc = {
                        "cash": f"現金模式，每次減少 {step} 元",
                        "ratio": f"比率模式，每次減少 {step*100}%",
                        "conditional": "條件模式，依價金範圍調整步幅",
                        "customize": "自訂模式，自動或手動配置步幅"
                    }.get(mode, f"{mode}模式")

                    # === 為每次滾算寫入紀錄單並創建對應工作表 ===
                    for i, (price, profit) in enumerate(zip(adjustment_record, profit_record)):
                        # 取得下一個可用的編號（用於滾算紀錄單和工作表）
                        record_number = self._get_next_record_number(wb)

                        # 取得對應的IRR結果
                        irr_data = irr_results[i] if i < len(irr_results) else {"project_irr": None, "equity_method_irr": None}

                        if i == 0:
                            remark = f"初始價金 - {mode_desc}，邊界值 {boundary} 元"
                        else:
                            reduction = adjustment_record[i-1] - price
                            remark = f"第{i}次滾算，減少 {reduction} 元 - {mode_desc}"

                        # 獲取利潤率
                        current_profit_rate = result.get("original_data", {}).get("profit_rate", 0.2)
                        if hasattr(self, 'current_profit_rate') and self.current_profit_rate is not None:
                            current_profit_rate = self.current_profit_rate

                        # 獲取Excel建置容量
                        capacity = result.get("original_data", {}).get("capacity", 436.1)

                        # 每kw信邦利潤：優先使用前端預計算值（與顯示一致），否則自行計算
                        precomputed_ppkw = irr_data.get('profit_per_kw')
                        if precomputed_ppkw is not None:
                            profit_per_kw = precomputed_ppkw
                        else:
                            profit_per_kw = price / (1 - current_profit_rate) - price
                        calculated_profit = profit_per_kw * capacity

                        # === 寫入滾算紀錄單（新增編號欄位） ===
                        sheet.cell(row=current_row, column=1, value=record_number)  # 編號
                        sheet.cell(row=current_row, column=2, value=price)  # 價金每KW

                        # IRR 以小數形式儲存（0.0175 = 1.75%），套 0.00% 格式
                        # 如此 XLSX.js 讀到 % 格式會用 v.m 顯示，不走 Math.round
                        def _irr_val(v):
                            return v / 100 if v is not None else None

                        irr_c3 = sheet.cell(row=current_row, column=3, value=_irr_val(irr_data.get("project_irr")))
                        irr_c3.number_format = '0.00%'
                        irr_c4 = sheet.cell(row=current_row, column=4, value=_irr_val(irr_data.get("cost_method_irr")))
                        irr_c4.number_format = '0.00%'
                        irr_c5 = sheet.cell(row=current_row, column=5, value=_irr_val(irr_data.get("equity_method_irr")))
                        irr_c5.number_format = '0.00%'

                        profit_rate_cell = sheet.cell(row=current_row, column=6, value=current_profit_rate)
                        profit_rate_cell.number_format = '0.00%'

                        sheet.cell(row=current_row, column=7, value=calculated_profit)  # 信邦利潤
                        sheet.cell(row=current_row, column=8, value=profit_per_kw)      # 每kw信邦利潤
                        sheet.cell(row=current_row, column=9, value=development_fee)   # 開發費
                        sheet.cell(row=current_row, column=10, value=remark)            # 備註

                        # === 創建對應編號的滾算紀錄工作表 ===
                        new_sheet_name = f"p{record_number}"
                        new_sheet = None
                        template_copied = False

                        # 方法1: 優先從輸入檔案中已有的模板工作表複製
                        for template_name in ["輸入公版(輸入模擬)", "輸入公版", "滾算紀錄1"]:
                            if template_name in wb.sheetnames:
                                new_sheet = wb.copy_worksheet(wb[template_name])
                                new_sheet.title = new_sheet_name
                                print(f"已從「{template_name}」複製為「{new_sheet_name}」(編號 {record_number})")
                                template_copied = True
                                break

                        # 方法2: 如果輸入檔案沒有模板，從外部模板檔案複製
                        if not template_copied:
                            template_file = os.path.join(self.excel_final_folder, "excel公版.xlsx")
                            if os.path.exists(template_file):
                                try:
                                    template_wb = load_workbook(template_file)
                                    # 嘗試從模板檔案中尋找合適的工作表
                                    for template_name in ["輸入公版(輸入模擬)", "輸入公版", "滾算紀錄1"]:
                                        if template_name in template_wb.sheetnames:
                                            template_sheet = template_wb[template_name]
                                            new_sheet = wb.create_sheet(new_sheet_name)

                                            # 複製所有單元格內容和格式
                                            for row in template_sheet.iter_rows():
                                                for cell in row:
                                                    new_cell = new_sheet[cell.coordinate]
                                                    new_cell.value = cell.value
                                                    if cell.has_style:
                                                        new_cell.font = cell.font.copy()
                                                        new_cell.border = cell.border.copy()
                                                        new_cell.fill = cell.fill.copy()
                                                        new_cell.number_format = cell.number_format
                                                        new_cell.protection = cell.protection.copy()
                                                        new_cell.alignment = cell.alignment.copy()

                                            # 複製列寬
                                            for col_letter, col_dim in template_sheet.column_dimensions.items():
                                                new_sheet.column_dimensions[col_letter].width = col_dim.width

                                            # 複製行高與隱藏狀態
                                            for row_num, row_dim in template_sheet.row_dimensions.items():
                                                new_sheet.row_dimensions[row_num].height = row_dim.height
                                                new_sheet.row_dimensions[row_num].hidden = row_dim.hidden

                                            template_wb.close()
                                            print(f"已從模板檔案「{template_name}」複製為「{new_sheet_name}」(編號 {record_number})")
                                            template_copied = True
                                            break

                                    if not template_copied:
                                        template_wb.close()
                                except Exception as e:
                                    print(f"從模板檔案複製失敗: {str(e)}")

                        # 方法3: 如果都沒有模板，創建空白工作表
                        if not template_copied:
                            new_sheet = wb.create_sheet(new_sheet_name)
                            print(f"已建立新工作表「{new_sheet_name}」(編號 {record_number}，無模板)")

                        # 填入該次滾算的數據
                        if new_sheet:
                            self._fill_record_sheet_with_data(new_sheet, result, i, record_number)

                        current_row += 1

                    # 添加空行分隔不同的滾算批次
                    current_row += 1

                # 為所有工作表自動調整欄寬
                for sheet_name in wb.sheetnames:
                    self._auto_adjust_column_widths(wb[sheet_name])

                wb.save(output_path)
                wb.close()
                _recalc(output_path)  # 填入公式快取，確保前端讀取正確數值
                return  # 成功完成，退出重試循環
                
            except PermissionError as e:
                if attempt < max_retries - 1:
                    print(f"檔案被鎖定，等待 {retry_delay} 秒後重試 (第 {attempt + 1}/{max_retries} 次)")
                    time.sleep(retry_delay)
                    retry_delay *= 2  # 指數退避
                else:
                    raise Exception(f"檔案被鎖定，無法寫入。請確保 Excel 檔案未被其他程式開啟: {str(e)}")
            except Exception as e:
                raise Exception(f"寫入 Excel 檔案失敗: {str(e)}")

    def execute_price_rolling(self,
                              mode: str = "cash",
                              boundary: int = 20000,
                              step: Union[int, float] = 1000,
                              profit_rate: float = None,
                              development_fee: int = None,
                              adjustment_times: int = 10,
                              sheet_name: str = None,
                              excel_file: str = None,
                              precomputed_irr_results: list = None,
                              **kwargs) -> Dict[str, Any]:
        """
        執行價金滾算

        Args:
            mode: 滾算模式 ("cash", "ratio", "conditional", "customize")
            boundary: 價金調整邊界
            step: 調整步幅（現金模式為整數，比率模式為小數）
            profit_rate: 信邦利潤率（如未指定，將從Excel檔案讀取）
            development_fee: 開發費用（如未指定，將從Excel檔案讀取）
            adjustment_times: 調整次數（自訂模式使用）
            sheet_name: Excel工作表名稱（用於IRR計算）
            excel_file: 【必需】Excel檔案路徑（絕對路徑或相對於Excel資料夾的檔案名）。對應當前聊天室使用的Excel檔案
            **kwargs: 其他模式特定參數

        Returns:
            執行結果
        """
        try:
            # 1. 讀取 Excel 檔案（必須指定）
            if excel_file is None:
                return {
                    "success": False,
                    "message": "必須指定Excel檔案路徑。請在調用時傳入 excel_file 參數，例如：excel_file='表 43_模擬輸入.xlsx'"
                }

            # 如果是相對路徑（只有檔案名），轉換為絕對路徑
            if not os.path.isabs(excel_file):
                excel_file = os.path.join(self.excel_folder, excel_file)

            # 檢查指定的檔案是否存在
            if not os.path.exists(excel_file):
                return {
                    "success": False,
                    "message": f"指定的Excel檔案不存在: {excel_file}"
                }

            print(f"使用Excel檔案: {excel_file}")

            # 2. 讀取數據
            data = self._read_excel_data(excel_file)
            initial_cost = int(data.get("equipment_cost", 30000))
            
            # 3. 使用Excel中的預設值，除非用戶明確指定
            if profit_rate is None:
                profit_rate = float(data.get("profit_rate", 0.1))
                print(f"使用Excel預設利潤率: {profit_rate}")
            else:
                print(f"使用用戶指定利潤率: {profit_rate}")
            
            # 設定當前利潤率，供Excel寫入時使用
            self.current_profit_rate = profit_rate
            
            if development_fee is None:
                development_fee = int(data.get("development_fee", 2000))
                print(f"使用Excel預設開發費: {development_fee}")
            else:
                print(f"使用用戶指定開發費: {development_fee}")

            # 4. 執行滾算計算
            results = []
            
            if mode == "cash":
                calculator = CashMode(boundary=boundary, step=int(step))
                adjustment_record = calculator.calculation(initial_cost)
                
            elif mode == "ratio":
                calculator = RatioMode(boundary=boundary, step=float(step))
                adjustment_record = calculator.calculation(initial_cost)
                
            elif mode == "conditional":
                maximum_value = kwargs.get("maximum_value", 50000)
                minimum_value = kwargs.get("minimum_value", 30000)
                condition_step_1 = kwargs.get("condition_step_1", 2000)
                condition_step_2 = kwargs.get("condition_step_2", 1000)
                condition_step_3 = kwargs.get("condition_step_3", 500)
                
                calculator = ConditionalMode(
                    boundary=boundary,
                    maximum_value=maximum_value,
                    minimum_value=minimum_value,
                    condition_step_1=condition_step_1,
                    condition_step_2=condition_step_2,
                    condition_step_3=condition_step_3
                )
                adjustment_record = calculator.calculation(initial_cost)
                
            elif mode == "customize":
                calculator = CustomizeMode(boundary=boundary, adjustment_times=adjustment_times)
                if kwargs.get("auto_config", True):
                    steps = calculator.automatic_configuration(initial_cost)
                else:
                    steps = kwargs.get("custom_steps", [1000] * adjustment_times)
                adjustment_record = calculator.calculation(steps, initial_cost)
                
            else:
                return {
                    "success": False,
                    "message": f"不支援的滾算模式: {mode}"
                }

            # 4. 計算成本結構調整
            cost_service = CostStructureService(profit_rate=profit_rate, development_fee=development_fee)
            cost_structure_adjusted = cost_service.equipment_cost_calculation(adjustment_record)
            profit_record = cost_service.get_profit(adjustment_record)
            
            # 5. 計算IRR（含 cash_flow_details，供填入現金流量表各列）
            # 無論是否有預計算結果，都必須執行完整計算以取得 cash_flow_details
            print("正在計算IRR（含金流明細）...")
            irr_results = self._calculate_irr_for_prices(excel_file, adjustment_record, profit_rate, development_fee, sheet_name)

            # 若有預計算 IRR，用其數值覆蓋（保持與前端顯示一致），但保留 cash_flow_details
            if precomputed_irr_results and len(precomputed_irr_results) == len(adjustment_record):
                print(f"套用預計算 IRR 數值，共 {len(precomputed_irr_results)} 筆")
                for i, precomputed in enumerate(precomputed_irr_results):
                    irr_results[i]['project_irr']     = precomputed.get('project_irr',     irr_results[i].get('project_irr'))
                    irr_results[i]['cost_method_irr'] = precomputed.get('cost_method_irr', irr_results[i].get('cost_method_irr'))
                    irr_results[i]['equity_method_irr'] = precomputed.get('equity_method_irr', irr_results[i].get('equity_method_irr'))
                    if precomputed.get('profit_per_kw') is not None:
                        irr_results[i]['profit_per_kw'] = precomputed['profit_per_kw']

            # 檢查是否有 IRR 計算失敗
            for irr_data in irr_results:
                if (irr_data.get('project_irr') is None or
                        irr_data.get('cost_method_irr') is None or
                        irr_data.get('equity_method_irr') is None):
                    return {"success": False, "message": "公版數值異常導致IRR計算錯誤"}

            # 6. 準備結果數據
            result = {
                "mode": mode,
                "initial_cost": initial_cost,
                "boundary": boundary,
                "step": step,
                "adjustment_record": adjustment_record,
                "cost_structure_adjusted": cost_structure_adjusted,
                "profit_record": profit_record,
                "irr_results": irr_results,
                "development_fee": development_fee,
                "original_data": data
            }
            results.append(result)

            # 7. 準備輸出檔案並寫入結果
            output_path = self._prepare_output_file(excel_file)
            self._write_results_to_excel(output_path, results)

            return {
                "success": True,
                "message": "價金滾算執行完成",
                "result": result,
                "output_file": output_path,
                "summary": {
                    "initial_cost": initial_cost,
                    "final_cost": adjustment_record[-1] if adjustment_record else initial_cost,
                    "adjustment_count": len(adjustment_record) - 1,
                    "total_reduction": initial_cost - (adjustment_record[-1] if adjustment_record else initial_cost)
                }
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"執行價金滾算時發生錯誤: {str(e)}"
            }


# 定義工具 Schema
EQUIPMENT_COST_TOOLS_SCHEMA = [
    {
        "type": "function",
        "function": {
            "name": "execute_price_rolling",
            "description": "【完整流程】執行設備成本價金滾算，包含完整流程：1) 從指定的Excel檔案讀取數據, 2) 執行價金滾算計算, 3) 計算每個價金的 IRR, 4) 將結果直接寫入原Excel檔案（新增工作表）。此工具會在原Excel檔案中累積滾算記錄，每次執行會新增「pX」（如 p1、p2）等工作表。如只需要計算分析不寫入檔案，請使用 calculate_price_rolling 工具。",
            "parameters": {
                "type": "object",
                "properties": {
                    "mode": {
                        "type": "string",
                        "enum": ["cash", "ratio", "conditional", "customize"],
                        "description": "滾算模式：cash(現金模式), ratio(比率模式), conditional(條件模式), customize(自訂模式)"
                    },
                    "boundary": {
                        "type": "integer",
                        "description": "價金調整的邊界值"
                    },
                    "step": {
                        "type": "number",
                        "description": "調整步幅，現金模式使用整數，比率模式使用 0-1 之間的小數"
                    },
                    "profit_rate": {
                        "type": "number",
                        "description": "信邦利潤率 (0-1 之間的小數)，如未指定將從Excel檔案讀取預設值"
                    },
                    "development_fee": {
                        "type": "integer",
                        "description": "開發費用，如未指定將從Excel檔案讀取預設值"
                    },
                    "adjustment_times": {
                        "type": "integer",
                        "description": "調整次數（自訂模式使用）"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "用於IRR計算的Excel工作表名稱（可選）"
                    },
                    "excel_file": {
                        "type": "string",
                        "description": "【必需】Excel檔案路徑（絕對路徑或相對於Excel資料夾的檔案名）。例如：'表 43_模擬輸入.xlsx' 或 'D:\\\\path\\\\to\\\\file.xlsx'。此參數對應當前聊天室使用的Excel檔案，必須明確指定，不會自動尋找。"
                    },
                    "maximum_value": {
                        "type": "integer",
                        "description": "條件模式：判斷價金範圍的最大值"
                    },
                    "minimum_value": {
                        "type": "integer",
                        "description": "條件模式：判斷價金範圍的最小值"
                    },
                    "condition_step_1": {
                        "type": "integer",
                        "description": "條件模式：價金 > maximum_value 時的 step"
                    },
                    "condition_step_2": {
                        "type": "integer",
                        "description": "條件模式：minimum_value <= 價金 <= maximum_value 時的 step"
                    },
                    "condition_step_3": {
                        "type": "integer",
                        "description": "條件模式：價金 < minimum_value 時的 step"
                    },
                    "auto_config": {
                        "type": "boolean",
                        "description": "自訂模式：是否自動配置 steps"
                    },
                    "custom_steps": {
                        "type": "array",
                        "items": {
                            "type": "integer"
                        },
                        "description": "自訂模式：手動指定的 steps 列表"
                    }
                },
                "required": ["mode", "excel_file"]
            }
        }
    }
]