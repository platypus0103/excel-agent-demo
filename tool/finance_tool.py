# tool/finance_tool.py
import numpy as np
import openpyxl
import numpy_financial as npf
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, List, Any, Optional


class FinanceTool:
    """財務計算工具類：用於太陽能發電專案的財務分析"""

    def __init__(self, excel_file: str = None, sheet_name: str = None):
        """
        初始化財務工具

        Args:
            excel_file: Excel 輸入文件路徑，如果為 None 則會自動搜尋 Excel 資料夾中的檔案
            sheet_name: 工作表名稱，如果為 None 則使用第一個工作表
        """
        self.excel_file = excel_file
        self.sheet_name = sheet_name
        self.data = None  # 儲存計算結果

    def set_excel_file(self, excel_file: str, sheet_name: str = None):
        """
        設定要使用的 Excel 檔案路徑和工作表

        Args:
            excel_file: Excel 檔案路徑
            sheet_name: 工作表名稱，如果為 None 則使用第一個工作表
        """
        self.excel_file = excel_file
        # 【重要】當更換檔案時，必須重設 sheet_name
        # 這樣才能自動使用新檔案的第一個工作表，而不是保留舊的工作表名稱
        self.sheet_name = None  # 如果是 None，則 _load_excel_data 會自動使用第一個工作表
        self.data = None  # 重置計算結果，因為檔案已更換

    def _round_decimal(self, value):
        """將 Decimal 四捨五入到整數"""
        if isinstance(value, Decimal):
            return int(value.quantize(Decimal('1'), rounding=ROUND_HALF_UP))
        return int(value)

    def _round_irr(self, value):
        """將 IRR 四捨五入到小數點後兩位"""
        if np.isnan(value):
            return None
        return round(value * 100, 2)

    def _read_irr_from_sheet(self, sheet_name: str) -> Optional[Dict]:
        """
        直接從工作表 B 欄搜尋 IRR 關鍵字，讀取 C 欄的預算值。
        適用於「滾算紀錄」等已有預算 IRR 的工作表。

        Returns:
            含 project_irr / cost_method_irr / equity_method_irr 的 dict，
            若任一值找不到則回傳 None（改用計算模式）
        """
        try:
            wb = openpyxl.load_workbook(self.excel_file, data_only=True)
            if sheet_name not in wb.sheetnames:
                return None
            ws = wb[sheet_name]

            irr_keywords = {
                "project_irr":      ["專案法", "Project IRR", "project irr", "專案IRR", "專案 IRR"],
                "cost_method_irr":  ["成本法", "Cost IRR", "cost irr", "成本IRR", "成本 IRR"],
                "equity_method_irr": ["權益法", "Equity IRR", "equity irr", "權益IRR", "權益 IRR"],
            }

            found = {}
            for row in ws.iter_rows():
                b_cell = row[1] if len(row) > 1 else None  # B 欄 (index 1)
                c_cell = row[2] if len(row) > 2 else None  # C 欄 (index 2)
                if b_cell is None or b_cell.value is None:
                    continue
                label = str(b_cell.value).strip()
                for key, keywords in irr_keywords.items():
                    if key not in found and any(kw in label for kw in keywords):
                        val = c_cell.value if c_cell else None
                        if val is not None:
                            try:
                                found[key] = round(float(val), 2)
                            except (ValueError, TypeError):
                                pass

            if len(found) == 3:
                return found
            return None
        except Exception:
            return None

    def _load_excel_data(self) -> Dict:
        """從 Excel 讀取數據"""
        try:
            # 如果沒有設定檔案路徑，嘗試自動搜尋
            if not self.excel_file:
                import os
                # 取得專案根目錄
                current_dir = os.path.dirname(os.path.abspath(__file__))
                parent_dir = os.path.dirname(current_dir)
                excel_dir = os.path.join(parent_dir, 'Excel User Data')

                # 搜尋 Excel 資料夾中的 .xlsx 或 .xls 檔案
                # 排除暫存檔（以 ~$ 開頭的檔案）
                if os.path.exists(excel_dir):
                    excel_files = [f for f in os.listdir(excel_dir) 
                                   if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
                    if excel_files:
                        self.excel_file = os.path.join(excel_dir, excel_files[0])
                        print(f"自動找到 Excel 檔案: {self.excel_file}")
                    else:
                        raise Exception("Excel 資料夾中沒有找到 Excel 檔案（排除暫存檔）")
                else:
                    raise Exception("Excel 資料夾不存在")

            wb = openpyxl.load_workbook(self.excel_file, data_only=True)

            # 如果沒有指定工作表名稱，使用第一個工作表
            if not self.sheet_name:
                self.sheet_name = wb.sheetnames[0]
                print(f"使用第一個工作表: {self.sheet_name}")

            # 檢查工作表是否存在
            if self.sheet_name not in wb.sheetnames:
                available_sheets = ', '.join(wb.sheetnames)
                raise Exception(f"找不到工作表 '{self.sheet_name}'。可用的工作表: {available_sheets}")

            s1 = wb[self.sheet_name]
            print(f"正在讀取工作表: {self.sheet_name}")

            def _si(cell_val, default):
                """安全轉整數：None / '-' / 空字串等非數值一律回傳預設值"""
                if cell_val is None:
                    return default
                try:
                    return int(float(str(cell_val).strip()))
                except (ValueError, TypeError):
                    return default

            def _sd(cell_val, default):
                """安全轉 Decimal：None / '-' / 空字串等非數值一律回傳預設值"""
                if cell_val is None:
                    return default
                try:
                    return Decimal(str(float(str(cell_val).strip())))
                except (ValueError, TypeError):
                    return default

            # 讀取並轉換所有必要數據
            data = {
                'project_name': s1['B2'].value,
                'year_start': _si(s1['B5'].value, None),
                'year_end': _si(s1['D5'].value, None),
                'capacities': _sd(s1['B4'].value, None),
                'electricity_generation': _sd(s1['C10'].value, None),
                'first_year_decade': _sd(s1['C11'].value, None),
                'every_year_decade': _sd(s1['C12'].value, None),
                'rate_of_sell': _sd(s1['C13'].value, None),
                'income_years': _si(s1['D10'].value, None),
                'equipment_cost': _sd(s1['C16'].value, None),
                'equipment_years': _si(s1['D16'].value, None),
                'profit_rate': _sd(s1['C17'].value, None),
                'development_cost': _sd(s1['C18'].value, None),
                'rent_mode': _si(s1['C19'].value, 1),
                'rent': _sd(s1['C20'].value, Decimal('0')),
                'rent_years': _si(s1['D20'].value, 0),
                'rent_method2_ratio': _sd(s1['C21'].value, Decimal('0')),
                'maintenance_cost': _sd(s1['C22'].value, Decimal('0')),
                'maintenance_years': _si(s1['D22'].value, 0),
                'insurance_cost': _sd(s1['C23'].value, Decimal('0')),
                'insurance_years': _si(s1['D23'].value, 0),
                'recycle_cost': _sd(s1['C24'].value, Decimal('0')),
                'recycle_years': _si(s1['D24'].value, 0),
                'interest_rate': _sd(s1['C25'].value, None),
                'income_tax': _sd(s1['C26'].value, None),
                'loan_rate': _sd(s1['C31'].value, None),
                'loan_amoritization_years': _si(s1['C32'].value, None),
                'dividend_rate': _sd(s1['C33'].value, None),
                'reduction_amoritization_years': _si(s1['C34'].value, None),
            }

            return data

        except Exception as e:
            raise Exception(f"讀取 Excel 失敗: {str(e)}")

    def _calculate_from_data(self, excel_data: Dict) -> Dict:
        """
        根據提供的數據執行財務計算

        Args:
            excel_data: 包含所有必要參數的字典

        Returns:
            計算結果字典
        """
        # 基本計算
        total_equipment_cost = ((excel_data['equipment_cost']/(Decimal('1')-excel_data['profit_rate']))+excel_data['development_cost'])*excel_data['capacities']
        year_equipment_cost = total_equipment_cost / Decimal(str(excel_data['equipment_years']))
        total_years = int(excel_data['year_end'] - excel_data['year_start'] + 1)

        # 計算發電收入
        first_year_electricity_generation = excel_data['electricity_generation'] * (Decimal('1')-excel_data['first_year_decade'])
        first_year_electricity_generation_income = first_year_electricity_generation * excel_data['rate_of_sell']
        temp = first_year_electricity_generation
        electricity_generation_income_list = [first_year_electricity_generation_income]
        electricity_generation=[first_year_electricity_generation]

        # 計算各項支出
        annual_rent = excel_data['rent']  # C20 直接為每年租金金額，不再攤提
        annual_maintenance = excel_data['maintenance_cost']*excel_data['capacities'] if excel_data['maintenance_years'] > 0 else Decimal('0')
        annual_insurance = excel_data['equipment_cost']*excel_data['capacities']*excel_data['insurance_cost'] if excel_data['insurance_years'] > 0 else Decimal('0')
        annual_recycle = excel_data['recycle_cost']*excel_data['capacities']/Decimal(str(excel_data['recycle_years'])) if excel_data['recycle_years'] > 0 else Decimal('0')

        # 模式 2 租金需要各年電費收入，先預建完整列表
        rent_mode = int(excel_data.get('rent_mode', 1))
        if rent_mode == 2:
            _t = first_year_electricity_generation
            _all_elec_income = [first_year_electricity_generation_income]
            for _yr in range(2, total_years + 1):
                _t = _t - (excel_data['electricity_generation'] * excel_data['every_year_decade'])
                _all_elec_income.append(_t * excel_data['rate_of_sell'])

        # 創建支出 lists
        rent_list = []
        maintenance_list = []
        insurance_list = []
        recycle_list = []
        depreciation_list = []

        for year in range(1, total_years + 1):
            # 租金
            if rent_mode == 2:
                ratio = excel_data.get('rent_method2_ratio', Decimal('0'))
                rent_list.append(_all_elec_income[year - 1] * ratio)
            elif year == 1:
                rent_list.append(Decimal('0'))
            else:
                rent_list.append(annual_rent)

            # 運維費
            if excel_data['maintenance_years'] and year <= excel_data['maintenance_years']:
                maintenance_list.append(annual_maintenance)
            else:
                maintenance_list.append(Decimal('0'))

            # 保險費
            if excel_data['insurance_years'] and year <= excel_data['insurance_years']:
                insurance_list.append(annual_insurance)
            else:
                insurance_list.append(Decimal('0'))

            # 回收費
            if excel_data['recycle_years'] and year <= excel_data['recycle_years']:
                recycle_list.append(annual_recycle)
            else:
                recycle_list.append(Decimal('0'))

            # 折舊
            if excel_data['equipment_years'] and year <= excel_data['equipment_years']:
                depreciation_list.append(year_equipment_cost)
            else:
                depreciation_list.append(Decimal('0'))

        # 計算利息和發電收入
        interest_list = []
        loan_amount = total_equipment_cost * excel_data['loan_rate']
        pay_back = loan_amount / Decimal(str(excel_data['loan_amoritization_years'])) if excel_data['loan_amoritization_years'] > 0 else Decimal('0')
        remaining_loan = loan_amount
        # pay_back_list: 長度為 total_years，超過貸款攤還年限後填 0
        pay_back_list = []
        for year in range(1, total_years + 1):
            if year <= excel_data['loan_amoritization_years']:
                pay_back_list.append(pay_back)
            else:
                pay_back_list.append(Decimal('0'))

        for year in range(1, total_years + 1):
            interest = remaining_loan * excel_data['interest_rate'] if remaining_loan > Decimal('0') else Decimal('0')
            interest_list.append(interest)

            remaining_loan -= pay_back
            if remaining_loan < Decimal('0'):
                remaining_loan = Decimal('0')

            if year > 1:
                temp = temp - (excel_data['electricity_generation'] * excel_data['every_year_decade'])
                electricity_generation_income_list.append(temp * excel_data['rate_of_sell'])
                electricity_generation.append(temp)

        # 計算現金流
        cash_flows = [-total_equipment_cost]
        total_cost = annual_rent + annual_maintenance + annual_insurance + annual_recycle

        first_year_cash_flow = first_year_electricity_generation_income - total_cost
        first_tax = (first_year_cash_flow - year_equipment_cost + annual_rent) * excel_data['income_tax']
        first_year_cash_flow -= first_tax
        cash_flows.append(first_year_cash_flow)

        remaining_loan = loan_amount
        remaining_loan -= pay_back

        for year in range(2, total_years + 1):
            following_years_electricity_generation_income = electricity_generation_income_list[year-1]
            annual_cash_flow = following_years_electricity_generation_income - rent_list[year-1] - maintenance_list[year-1] - insurance_list[year-1] - recycle_list[year-1]
            annual_tax = (annual_cash_flow - year_equipment_cost) * excel_data['income_tax']
            annual_cash_flow -= annual_tax
            cash_flows.append(annual_cash_flow)

            remaining_loan -= pay_back
            if remaining_loan < Decimal('0'):
                remaining_loan = Decimal('0')

        # 計算稅後淨利
        net_profit_after_tax = []
        remaining_loan_list = loan_amount

        for year in range(1, total_years + 1):
            incom = electricity_generation_income_list[year-1]
            rent_expense = rent_list[year-1]
            maintenance_expense = maintenance_list[year-1]
            insurance_expense = insurance_list[year-1]
            recycle_expense = recycle_list[year-1]
            interest_expense = interest_list[year-1]
            depreciation_expense = depreciation_list[year-1]

            if year == 1:
                net_profit_after_tax_year = incom - maintenance_expense - insurance_expense - recycle_expense - interest_expense - depreciation_expense
            else:
                net_profit_after_tax_year = incom - rent_expense - maintenance_expense - insurance_expense - recycle_expense - interest_expense - depreciation_expense

            tax_flow = net_profit_after_tax_year * excel_data['income_tax']
            net_profit_after_tax_year -= tax_flow
            net_profit_after_tax.append(net_profit_after_tax_year)

        # 計算 IRR
        irr = npf.irr([float(cf) for cf in cash_flows])

        # 成本法 IRR
        cost_method_initial_investment = (total_equipment_cost*(Decimal('1') - excel_data['loan_rate']))
        cost_method_cash_flows = [-cost_method_initial_investment]
        reduction_at_the_end = (cost_method_initial_investment/Decimal(str(excel_data['reduction_amoritization_years'])))
        dividend_list=np.zeros(total_years)
        for year in range(-1,-excel_data['reduction_amoritization_years']-1,-1):#用來填值 不是用來計算
            dividend_list[year]=reduction_at_the_end
        for year in range(1, total_years):
            dividend = net_profit_after_tax[year-1]*excel_data['dividend_rate']
            cost_method_cash_flows.append(dividend)

        for year in range(total_years - excel_data['reduction_amoritization_years'], total_years):
            cost_method_cash_flows[year] += reduction_at_the_end

        last_cost_method_cash_flows = sum(net_profit_after_tax)-(sum(net_profit_after_tax)-net_profit_after_tax[-1])*excel_data['dividend_rate']
        cost_method_cash_flows.append(last_cost_method_cash_flows)
        cost_method_irr = npf.irr([float(cf) for cf in cost_method_cash_flows])

        # 權益法 IRR
        equity_method_initial_investment = -(total_equipment_cost * (Decimal('1') - excel_data['loan_rate']))+net_profit_after_tax[0]
        equity_method_cash_flows = [equity_method_initial_investment]

        for year in range(2, total_years + 1):
            equity_method_cash_flows.append(net_profit_after_tax[year-1])
            if (year-1) >= (total_years - excel_data['reduction_amoritization_years']):
                equity_method_cash_flows[year-1] += reduction_at_the_end

        equity_method_irr = npf.irr([float(cf) for cf in equity_method_cash_flows])

        return {
            'project_name': excel_data['project_name'],
            'year_start': excel_data['year_start'],
            'year_end': excel_data['year_end'],
            'total_years': total_years,
            'total_equipment_cost': total_equipment_cost,
            'loan_amount': loan_amount,
            'pay_back': pay_back,
            'electricity_generation_income_list': electricity_generation_income_list,
            'rent_list': rent_list,
            'maintenance_list': maintenance_list,
            'insurance_list': insurance_list,
            'recycle_list': recycle_list,
            'interest_list': interest_list,
            'depreciation_list': depreciation_list,
            'pay_back_list': pay_back_list,  # 貸款還款列表
            'dividend_list': dividend_list,  # 年底減資列表
            'cash_flows': cash_flows,
            'net_profit_after_tax': net_profit_after_tax,
            'cost_method_cash_flows': cost_method_cash_flows,
            'equity_method_cash_flows': equity_method_cash_flows,
            'project_irr': irr,
            'cost_method_irr': cost_method_irr,
            'equity_method_irr': equity_method_irr,
            'raw_equipment_cost': excel_data['equipment_cost'],
            'profit_rate': excel_data['profit_rate']
        }

    def _calculate_all(self):
        """執行所有財務計算"""
        if self.data is not None:
            return  # 已經計算過了

        # 讀取 Excel 數據
        excel_data = self._load_excel_data()
        
        # 執行計算
        self.data = self._calculate_from_data(excel_data)

    def calculate_scenario_irr(
        self, 
        equipment_cost: int, 
        profit_rate: float = None, 
        development_cost: int = None,
        sheet_name: str = None
    ) -> Dict:
        """
        計算特定價金下的 IRR（不修改 Excel 檔案）

        Args:
            equipment_cost: 模擬的價金 / kW
            profit_rate: 利潤率 (可選，若提供則覆蓋 Excel 設定)
            development_cost: 開發費 (可選，若提供則覆蓋 Excel 設定)
            sheet_name: 工作表名稱

        Returns:
            包含三種 IRR 的字典
        """
        try:
            # 如果指定了工作表名稱，臨時設定
            original_sheet = self.sheet_name
            if sheet_name:
                self.sheet_name = sheet_name
            
            # 讀取原始數據
            excel_data = self._load_excel_data()
            
            # 覆蓋價金
            excel_data['equipment_cost'] = Decimal(str(equipment_cost))
            
            # 覆蓋其他參數 (如果有提供)
            if profit_rate is not None:
                excel_data['profit_rate'] = Decimal(str(profit_rate))
            
            if development_cost is not None:
                excel_data['development_cost'] = Decimal(str(development_cost))
            
            # 執行計算
            result = self._calculate_from_data(excel_data)

            # 恢復工作表設定
            if sheet_name and original_sheet != sheet_name:
                self.sheet_name = original_sheet

            return {
                "project_irr": self._round_irr(result['project_irr']),
                "cost_method_irr": self._round_irr(result['cost_method_irr']),
                "equity_method_irr": self._round_irr(result['equity_method_irr']),
                # 金流明細資料（每年）
                "cash_flow_details": {
                    "rent_list": [float(x) for x in result['rent_list']],
                    "maintenance_list": [float(x) for x in result['maintenance_list']],
                    "insurance_list": [float(x) for x in result['insurance_list']],
                    "recycle_list": [float(x) for x in result['recycle_list']],
                    "interest_list": [float(x) for x in result['interest_list']],
                    "depreciation_list": [float(x) for x in result['depreciation_list']],
                    "pay_back_list": [float(x) for x in result['pay_back_list']],  # 貸款還款
                    "dividend_list": [float(x) for x in result['dividend_list']],  # 年底減資
                    "electricity_generation_income_list": [float(x) for x in result['electricity_generation_income_list']],
                    "net_profit_after_tax": [float(x) for x in result['net_profit_after_tax']],
                    "cash_flows": [float(x) for x in result['cash_flows']],
                    "cost_method_cash_flows": [float(x) for x in result['cost_method_cash_flows']],
                    "equity_method_cash_flows": [float(x) for x in result['equity_method_cash_flows']],
                    "total_years": result['total_years']
                }
            }
        except Exception as e:
            print(f"⚠️ calculate_scenario_irr 錯誤 (equipment_cost={equipment_cost}): {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                "project_irr": None,
                "cost_method_irr": None,
                "equity_method_irr": None,
                "error": str(e)
            }

    def calculate_irr(self, sheet_name: str = None) -> Dict:
        """
        計算專案的 IRR（內部報酬率）

        Args:
            sheet_name: 工作表名稱，如果不指定則使用當前設定的工作表或第一個工作表

        Returns:
            包含三種 IRR 的結果字典
        """
        try:
            # 如果指定了工作表名稱，臨時設定
            original_sheet = self.sheet_name
            target_sheet = sheet_name if sheet_name else self.sheet_name

            # 優先：從工作表 B/C 欄直接讀取預算 IRR（適用於滾算紀錄等已計算工作表）
            direct = self._read_irr_from_sheet(target_sheet) if target_sheet else None
            if direct:
                return {
                    "success": True,
                    "message": f"IRR 讀取完成 (工作表: {target_sheet})",
                    "data": {
                        "sheet_name": target_sheet,
                        "project_irr": direct["project_irr"],
                        "cost_method_irr": direct["cost_method_irr"],
                        "equity_method_irr": direct["equity_method_irr"],
                    }
                }

            # 備用：從參數重新計算
            if sheet_name:
                self.sheet_name = sheet_name
                self.data = None

            self._calculate_all()

            result = {
                "success": True,
                "message": f"IRR 計算完成 (工作表: {self.sheet_name})",
                "data": {
                    "project_name": self.data['project_name'],
                    "sheet_name": self.sheet_name,
                    "project_irr": self._round_irr(self.data['project_irr']),
                    "cost_method_irr": self._round_irr(self.data['cost_method_irr']),
                    "equity_method_irr": self._round_irr(self.data['equity_method_irr']),
                    "raw_equipment_cost": self._round_decimal(self.data['raw_equipment_cost']),
                    "profit_rate": float(self.data['profit_rate'])
                }
            }

            # 如果臨時修改了工作表，恢復原設定
            if sheet_name and original_sheet != sheet_name:
                self.sheet_name = original_sheet
                self.data = None

            return result
        except Exception as e:
            return {
                "success": False,
                "message": f"IRR 計算失敗: {str(e)}"
            }

    def get_cash_flow(self, year: Optional[int] = None, sheet_name: str = None) -> Dict:
        """
        獲取現金流資料

        Args:
            year: 指定年份（可選），如果不指定則返回所有年份
            sheet_name: 工作表名稱，如果不指定則使用當前設定的工作表或第一個工作表

        Returns:
            現金流資料字典
        """
        try:
            # 如果指定了工作表名稱，臨時設定
            original_sheet = self.sheet_name
            if sheet_name:
                self.sheet_name = sheet_name
                self.data = None

            self._calculate_all()

            if year is not None:
                if year < 0 or year > self.data['total_years']:
                    return {
                        "success": False,
                        "message": f"年份 {year} 超出範圍 (0-{self.data['total_years']})"
                    }

                result = {
                    "success": True,
                    "message": f"第 {year} 年現金流 (工作表: {self.sheet_name})",
                    "data": {
                        "year": year,
                        "sheet_name": self.sheet_name,
                        "cash_flow": self._round_decimal(self.data['cash_flows'][year])
                    }
                }
            else:
                cash_flows_data = []
                for i, cf in enumerate(self.data['cash_flows']):
                    cash_flows_data.append({
                        "year": i,
                        "cash_flow": self._round_decimal(cf)
                    })

                result = {
                    "success": True,
                    "message": f"所有年份現金流 (工作表: {self.sheet_name})",
                    "data": {
                        "sheet_name": self.sheet_name,
                        "total_years": self.data['total_years'],
                        "cash_flows": cash_flows_data
                    }
                }

            # 如果臨時修改了工作表，恢復原設定
            if sheet_name and original_sheet != sheet_name:
                self.sheet_name = original_sheet
                self.data = None

            return result
        except Exception as e:
            return {
                "success": False,
                "message": f"獲取現金流失敗: {str(e)}"
            }

    def get_net_profit(self, year: Optional[int] = None, sheet_name: str = None) -> Dict:
        """
        獲取稅後淨利資料

        Args:
            year: 指定年份（可選），如果不指定則返回所有年份
            sheet_name: 工作表名稱，如果不指定則使用當前設定的工作表或第一個工作表

        Returns:
            稅後淨利資料字典
        """
        try:
            # 如果指定了工作表名稱，臨時設定
            original_sheet = self.sheet_name
            if sheet_name:
                self.sheet_name = sheet_name
                self.data = None

            self._calculate_all()

            if year is not None:
                if year < 1 or year > self.data['total_years']:
                    return {
                        "success": False,
                        "message": f"年份 {year} 超出範圍 (1-{self.data['total_years']})"
                    }

                result = {
                    "success": True,
                    "message": f"第 {year} 年稅後淨利 (工作表: {self.sheet_name})",
                    "data": {
                        "year": year,
                        "sheet_name": self.sheet_name,
                        "net_profit_after_tax": self._round_decimal(self.data['net_profit_after_tax'][year-1])
                    }
                }
            else:
                net_profit_data = []
                for i, np in enumerate(self.data['net_profit_after_tax']):
                    net_profit_data.append({
                        "year": i + 1,
                        "net_profit_after_tax": self._round_decimal(np)
                    })

                result = {
                    "success": True,
                    "message": f"所有年份稅後淨利 (工作表: {self.sheet_name})",
                    "data": {
                        "sheet_name": self.sheet_name,
                        "total_years": self.data['total_years'],
                        "net_profit_after_tax": net_profit_data
                    }
                }

            # 如果臨時修改了工作表，恢復原設定
            if sheet_name and original_sheet != sheet_name:
                self.sheet_name = original_sheet
                self.data = None

            return result
        except Exception as e:
            return {
                "success": False,
                "message": f"獲取稅後淨利失敗: {str(e)}"
            }

    def get_year_detail(self, year: int, sheet_name: str = None) -> Dict:
        """
        獲取指定年份的詳細財務資料

        Args:
            year: 年份（1-total_years）
            sheet_name: 工作表名稱，如果不指定則使用當前設定的工作表或第一個工作表

        Returns:
            該年份的詳細財務資料字典
        """
        try:
            # 如果指定了工作表名稱，臨時設定
            original_sheet = self.sheet_name
            if sheet_name:
                self.sheet_name = sheet_name
                self.data = None

            self._calculate_all()

            if year < 1 or year > self.data['total_years']:
                return {
                    "success": False,
                    "message": f"年份 {year} 超出範圍 (1-{self.data['total_years']})"
                }

            idx = year - 1

            result = {
                "success": True,
                "message": f"第 {year} 年詳細資料 (工作表: {self.sheet_name})",
                "data": {
                    "year": year,
                    "sheet_name": self.sheet_name,
                    "electricity_generation_income": self._round_decimal(self.data['electricity_generation_income_list'][idx]),
                    "rent": self._round_decimal(self.data['rent_list'][idx]),
                    "maintenance": self._round_decimal(self.data['maintenance_list'][idx]),
                    "insurance": self._round_decimal(self.data['insurance_list'][idx]),
                    "recycle": self._round_decimal(self.data['recycle_list'][idx]),
                    "interest": self._round_decimal(self.data['interest_list'][idx]),
                    "depreciation": self._round_decimal(self.data['depreciation_list'][idx]),
                    "cash_flow": self._round_decimal(self.data['cash_flows'][year]),
                    "net_profit_after_tax": self._round_decimal(self.data['net_profit_after_tax'][idx])
                }
            }

            # 如果臨時修改了工作表，恢復原設定
            if sheet_name and original_sheet != sheet_name:
                self.sheet_name = original_sheet
                self.data = None

            return result
        except Exception as e:
            return {
                "success": False,
                "message": f"獲取年度詳細資料失敗: {str(e)}"
            }

    def get_project_summary(self, sheet_name: str = None) -> Dict:
        """
        獲取專案摘要資訊

        Args:
            sheet_name: 工作表名稱，如果不指定則使用當前設定的工作表或第一個工作表

        Returns:
            專案摘要資訊字典
        """
        try:
            # 如果指定了工作表名稱，臨時設定
            original_sheet = self.sheet_name
            if sheet_name:
                self.sheet_name = sheet_name
                self.data = None

            self._calculate_all()

            result = {
                "success": True,
                "message": f"專案摘要 (工作表: {self.sheet_name})",
                "data": {
                    "project_name": self.data['project_name'],
                    "sheet_name": self.sheet_name,
                    "year_start": self.data['year_start'],
                    "year_end": self.data['year_end'],
                    "total_years": self.data['total_years'],
                    "total_equipment_cost": self._round_decimal(self.data['total_equipment_cost']),
                    "loan_amount": self._round_decimal(self.data['loan_amount']),
                    "pay_back": self._round_decimal(self.data['pay_back']),
                    "project_irr": self._round_irr(self.data['project_irr']),
                    "cost_method_irr": self._round_irr(self.data['cost_method_irr']),
                    "equity_method_irr": self._round_irr(self.data['equity_method_irr']),
                    "total_cash_flow": self._round_decimal(sum(self.data['cash_flows'])),
                    "raw_equipment_cost": self._round_decimal(self.data['raw_equipment_cost']),
                    "profit_rate": float(self.data['profit_rate'])
                }
            }

            # 如果臨時修改了工作表，恢復原設定
            if sheet_name and original_sheet != sheet_name:
                self.sheet_name = original_sheet
                self.data = None

            return result
        except Exception as e:
            return {
                "success": False,
                "message": f"獲取專案摘要失敗: {str(e)}"
            }


# 定義工具的 schema（用於 function calling）
FINANCE_TOOLS_SCHEMA = [
    {
        "type": "function",
        "function": {
            "name": "calculate_irr",
            "description": "計算太陽能發電專案的內部報酬率（IRR），包括專案法IRR、成本法IRR和權益法IRR。可以指定要計算的工作表（sheet），如果不指定則使用第一個工作表。",
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet_name": {
                        "type": "string",
                        "description": "Excel 工作表名稱，例如：'模擬輸入'、'Sheet1'、'輸入公版(輸入模擬)'。如果用戶提到特定工作表名稱，請使用該名稱。不指定則使用第一個工作表。"
                    }
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_cash_flow",
            "description": "獲取專案的現金流資料。可以查詢特定年份或所有年份的現金流，也可以指定要查詢的工作表。",
            "parameters": {
                "type": "object",
                "properties": {
                    "year": {
                        "type": "integer",
                        "description": "年份（0表示初始投資，1-N表示各年份）。不指定則返回所有年份"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Excel 工作表名稱。如果用戶提到特定工作表名稱，請使用該名稱。不指定則使用第一個工作表。"
                    }
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_net_profit",
            "description": "獲取專案的稅後淨利資料。可以查詢特定年份或所有年份的稅後淨利，也可以指定要查詢的工作表。",
            "parameters": {
                "type": "object",
                "properties": {
                    "year": {
                        "type": "integer",
                        "description": "年份（1-N）。不指定則返回所有年份"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Excel 工作表名稱。如果用戶提到特定工作表名稱，請使用該名稱。不指定則使用第一個工作表。"
                    }
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_year_detail",
            "description": "獲取指定年份的詳細財務資料，包括發電收入、各項支出、現金流和稅後淨利。可以指定要查詢的工作表。",
            "parameters": {
                "type": "object",
                "properties": {
                    "year": {
                        "type": "integer",
                        "description": "年份（1-N）"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Excel 工作表名稱。如果用戶提到特定工作表名稱，請使用該名稱。不指定則使用第一個工作表。"
                    }
                },
                "required": ["year"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_project_summary",
            "description": "獲取專案的整體摘要資訊，包括專案名稱、年限、總投資、IRR等關鍵指標。可以指定要查詢的工作表。",
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet_name": {
                        "type": "string",
                        "description": "Excel 工作表名稱。如果用戶提到特定工作表名稱，請使用該名稱。不指定則使用第一個工作表。"
                    }
                },
                "required": []
            }
        }
    }
]
