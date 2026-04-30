"""
AI Agent 藍圖 - 處理聊天請求
"""
from flask import Blueprint, request, jsonify, session, send_file
from werkzeug.utils import secure_filename
import sys
import os
import re

# 調整匯入路徑
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

from utils.app_logger import log_action, log_error


def _get_excel_dir():
    """根據登入用戶回傳用戶專屬的 Excel 目錄（以 email 命名），若未登入則使用根目錄"""
    base = os.path.join(parent_dir, 'Excel User Data')
    user_email = session.get('user_email')
    d = os.path.join(base, user_email) if user_email else base
    os.makedirs(d, exist_ok=True)
    return d


def _find_excel_file(case_id='', case_name='', original_filename=''):
    """
    通用 Excel 檔案定址，優先順序：
    1. case_id + original_filename  → {case_id}_{original_filename}（最精準）
    2. case_name + original_filename → {case_name}_{original_filename}（相容舊檔）
    3. case_id 前綴搜尋              → {case_id}_*.xlsx
    4. case_name 前綴搜尋            → {case_name}_*.xlsx
    5. original_filename 包含搜尋
    6. 目錄內任意 xlsx               → 最後保底
    """
    excel_dir = _get_excel_dir()
    all_files = [f for f in os.listdir(excel_dir)
                 if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')] if os.path.exists(excel_dir) else []

    # 方法1: case_id + original_filename
    if case_id and original_filename:
        candidate = os.path.join(excel_dir, f"{case_id}_{original_filename}")
        if os.path.exists(candidate):
            return candidate

    # 方法2: case_name + original_filename
    if case_name and original_filename:
        candidate = os.path.join(excel_dir, f"{case_name}_{original_filename}")
        if os.path.exists(candidate):
            return candidate

    # 方法3: case_id 前綴
    if case_id:
        matches = [f for f in all_files if f.startswith(f"{case_id}_")]
        if matches:
            return os.path.join(excel_dir, matches[0])

    # 方法4: case_name 前綴
    if case_name:
        matches = [f for f in all_files if f.startswith(f"{case_name}_")]
        if matches:
            return os.path.join(excel_dir, matches[0])

    # 方法5: original_filename 包含搜尋
    if original_filename:
        matches = [f for f in all_files if original_filename in f]
        if matches:
            return os.path.join(excel_dir, matches[0])

    # 方法6: 保底，取第一個 xlsx
    if all_files:
        return os.path.join(excel_dir, all_files[0])

    return None

try:
    from core.agent import AIAgent
    from config.settings import DEFAULT_CONFIG
except ImportError as e:
    print(f"嚴重警告: 無法匯入 AI Agent 核心模組: {e}")
    AIAgent = None
    DEFAULT_CONFIG = None

agent_bp = Blueprint('agent_bp', __name__)

# ── 使用說明文字 ──────────────────────────────────────────────
_HELP_TEXT = """## Excel AI 助手 使用說明

歡迎使用 Excel AI 助手！以下是主要功能的操作方式。

---

### 一、建立案場與上傳 Excel
1. 點擊左側「＋ 新增案場」建立案場
2. 在案場中點擊「上傳 Excel」，選擇試算表檔案（.xlsx）
3. 上傳後系統會自動載入試算表

---

### 二、查詢數據
直接用自然語言詢問，例如：
- `A3 是多少？`
- `查詢 2025 年的保險費`
- `p2 的現金流是多少`

---

### 三、修改 Excel 數值
直接說明要改什麼，例如：
- `把保險費 2025~2030 改成 -50000`
- `修改第 48 行 2026 年改成 -40000`
- `設備費用改成 28000`

系統會顯示確認訊息，回覆「**是**」後執行修改。

---

### 支援的關鍵字
輸入以下任一關鍵字可重新顯示此說明：
`使用說明`、`說明`、`help`、`如何使用`、`怎麼用`
"""

# 全域 Agent 實例
_agent_instance = None

def get_agent():
    """獲取或創建 AI Agent 實例"""
    global _agent_instance
    if _agent_instance is None:
        if AIAgent is None or DEFAULT_CONFIG is None:
            raise RuntimeError("AI Agent 核心模組未載入")

        print("初始化 AI Agent...")
        _agent_instance = AIAgent(DEFAULT_CONFIG)
        print("AI Agent 初始化完成")

    return _agent_instance






@agent_bp.route('/agent_chat', methods=['POST'])
def agent_chat():
    """
    處理來自前端的 Agent 聊天請求
    """
    if not request.is_json:
        return jsonify({"error": "請求必須是 JSON 格式"}), 400

    data = request.json
    user_query = data.get('query')

    # 獲取案場資訊
    case_name = data.get('case_name', '')
    case_id = str(data.get('case_id', ''))
    original_filename = data.get('original_filename', '')
    sheet_name = data.get('sheet_name', None)  # 工作表名稱，可選

    if not user_query:
        return jsonify({"error": "請提供 'query' 參數"}), 400

    # ── 千分位逗號正規化：-40,000 → -40000（避免 regex 解析與 LLM tool call 錯誤）──
    user_query = re.sub(
        r'(?<!\d)(-?\d{1,3}(?:,\d{3})+)(?!\d)',
        lambda m: m.group(0).replace(',', ''),
        user_query
    )

    # ── 使用說明快捷回應（不走 AI，直接回傳固定說明）──
    _HELP_KEYWORDS = {'使用說明', '說明', 'help', '幫助', '如何使用', '怎麼用', '功能介紹', '教學'}
    if user_query.strip().lower() in _HELP_KEYWORDS or user_query.strip() in _HELP_KEYWORDS:
        return jsonify({
            "query": user_query,
            "response": _HELP_TEXT,
            "excel_modified": False
        })

    print(f"\n--- API 請求: 執行 Agent 對話 ---")
    print(f"收到查詢: {user_query}")
    print(f"案場資訊: case_id={case_id}, case_name={case_name}, original_filename={original_filename}, sheet_name={sheet_name}")

    user_email = session.get('user_email', 'anonymous')
    log_action(user_email, 'agent_chat',
               f"case={case_name}({case_id}) query={user_query[:200]}")

    try:
        # 獲取 Agent 實例
        agent = get_agent()

        # 使用通用定址 helper 取得 Excel 檔案路徑
        excel_path = _find_excel_file(case_id=case_id, case_name=case_name, original_filename=original_filename)
        if excel_path:
            agent.tool_manager.set_excel_file(excel_path, sheet_name)
            print(f"已設定財務工具 Excel 檔案: {excel_path}")
        else:
            print(f"警告: 找不到 Excel 檔案 (case_id={case_id}, case_name={case_name})")

        # ── 取得可用工作表清單，注入到 query，避免模型幻覺 ──
        sheet_names_prefix = ''
        current_excel_path = excel_path
        if current_excel_path and os.path.exists(current_excel_path):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(current_excel_path, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
                sheet_names_prefix = f"[目前 Excel 可用工作表: {', '.join(sheet_names)}]\n"
            except Exception:
                pass

        # ── LLM function calling ──
        print("交由 LLM function calling 處理...")
        current_sheet_info = f"[使用者目前觀看的工作表：{sheet_name}]\n" if sheet_name else ""
        enhanced_query = f"{sheet_names_prefix}{current_sheet_info}{user_query}"
        agent_response = agent.chat(enhanced_query)
        excel_tools = ['write_excel_cell', 'delete_excel_cell', 'edit_sheet_by_field']
        excel_modified = any(tool in agent.last_used_tools for tool in excel_tools)

        print(f"Agent 回應成功 (Excel modified: {excel_modified})")
        log_action(user_email, 'agent_chat_done',
                   f"case={case_name}({case_id}) excel_modified={excel_modified} "
                   f"tools_used={agent.last_used_tools} response_len={len(agent_response)}")
        return jsonify({
            "query": user_query,
            "response": agent_response,
            "excel_modified": excel_modified
        })

    except Exception as e:
        import traceback
        print(f"AI Agent 處理錯誤: {e}")
        print(traceback.format_exc())
        log_error(user_email, 'agent_chat',
                  e, f"case={case_name}({case_id}) query={user_query[:200]}")
        return jsonify({"error": "系統處理時發生問題，請稍後再試。"}), 500

@agent_bp.route('/download_excel', methods=['GET'])
def download_excel():
    try:
        case_id = request.args.get('case_id', '')
        case_name = request.args.get('case_name', '')
        original_filename = request.args.get('original_filename', '')

        excel_path = _find_excel_file(case_id=case_id, case_name=case_name, original_filename=original_filename)

        if not excel_path or not os.path.exists(excel_path):
            return jsonify({"status": "error", "error": f"找不到案場「{case_name}」的 Excel 檔案"}), 404

        log_action(session.get('user_email', 'anonymous'), 'download_excel',
                   f"case={case_name}({case_id}) file={os.path.basename(excel_path)}")
        return send_file(
            excel_path,
            as_attachment=True,
            download_name=os.path.basename(excel_path),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        import traceback
        log_error(session.get('user_email', 'anonymous'), 'download_excel', e, f"case={case_name}({case_id})")
        return jsonify({"status": "error", "error": f"下載失敗: {str(e)}"}), 500


AVAILABLE_MODELS = ['qwen3:4b', 'qwen3:14b', 'qwen3:32b']

@agent_bp.route('/agent/model', methods=['GET'])
def get_model():
    """取得目前使用的模型"""
    try:
        agent = get_agent()
        return jsonify({"status": "success", "model": agent.config.model_name})
    except Exception as e:
        log_error(session.get('user_email', 'anonymous'), 'get_model', e)
        return jsonify({"status": "error", "error": str(e)}), 500

@agent_bp.route('/agent/model', methods=['POST'])
def set_model():
    """切換模型"""
    try:
        data = request.get_json()
        model = data.get('model', '')
        if model not in AVAILABLE_MODELS:
            return jsonify({"status": "error", "error": f"不支援的模型，可選: {AVAILABLE_MODELS}"}), 400
        agent = get_agent()
        agent.config.model_name = model
        agent.connection.model_name = model
        log_action(session.get('user_email', 'anonymous'), 'set_model', f"model={model}")
        print(f"[模型切換] → {model}")
        return jsonify({"status": "success", "model": model})
    except Exception as e:
        log_error(session.get('user_email', 'anonymous'), 'set_model', e, f"model={data.get('model', '') if 'data' in dir() else ''}")
        return jsonify({"status": "error", "error": str(e)}), 500

@agent_bp.route('/release_memory', methods=['POST'])
def release_memory():
    """釋放 Ollama 模型記憶體（keep_alive=0）"""
    try:
        import ollama
        agent = get_agent()
        model_name = agent.config.model_name
        ollama.generate(model=model_name, keep_alive=0)
        log_action(session.get('user_email', 'anonymous'), 'release_memory', f"model={model_name}")
        return jsonify({"status": "success", "model": model_name})
    except Exception as e:
        log_error(session.get('user_email', 'anonymous'), 'release_memory', e)
        return jsonify({"status": "error", "error": str(e)}), 500

@agent_bp.route('/reset', methods=['POST'])
def reset_agent():
    """重置 Agent 對話歷史"""
    try:
        agent = get_agent()
        agent.reset(keep_system=True)
        log_action(session.get('user_email', 'anonymous'), 'reset_agent', '')
        return jsonify({
            "status": "success",
            "message": "對話已重置"
        })
    except Exception as e:
        log_error(session.get('user_email', 'anonymous'), 'reset_agent', e)
        return jsonify({"error": str(e)}), 500

@agent_bp.route('/history', methods=['GET'])
def get_history():
    """獲取對話歷史"""
    try:
        agent = get_agent()
        messages = agent.conversation.get_messages()
        return jsonify({
            "status": "success",
            "history": messages
        })
    except Exception as e:
        log_error(session.get('user_email', 'anonymous'), 'get_history', e)
        return jsonify({"error": str(e)}), 500

@agent_bp.route('/upload_excel', methods=['POST'])
def upload_excel():
    """
    上傳 Excel 檔案到後端 Excel 資料夾
    """
    try:
        # 檢查是否有檔案
        if 'file' not in request.files:
            return jsonify({"error": "沒有檔案"}), 400

        file = request.files['file']

        # 檢查檔案名稱
        if file.filename == '':
            return jsonify({"error": "檔案名稱為空"}), 400

        # 獲取原始檔名（從前端傳來的參數）
        original_filename = request.form.get('original_filename', file.filename)

        # 檢查檔案類型
        allowed_extensions = {'.xlsx', '.xls'}
        file_ext = os.path.splitext(original_filename)[1].lower()
        if file_ext not in allowed_extensions:
            return jsonify({"error": f"不支援的檔案格式: {file_ext}"}), 400

        # 獲取案場 ID 和前端顯示名稱
        case_id = request.form.get('case_id', '')
        case_name = request.form.get('case_name', '')  # 前端顯示名稱（例如：表 1、財報）

        # 確保 Excel 目錄存在
        excel_dir = _get_excel_dir()
        if not os.path.exists(excel_dir):
            os.makedirs(excel_dir)
            print(f"創建 Excel 目錄: {excel_dir}")

        # 優先用 case_id 作為前綴（永久唯一，不受改名影響）
        if case_id:
            filename = f"{case_id}_{original_filename}"  # 例如：1_公版.xlsx
        else:
            filename = f"{case_name}_{original_filename}"  # fallback
        file_path = os.path.join(excel_dir, filename)
        file.save(file_path)
        print(f"Excel 檔案已儲存: {file_path}")
        print(f"檔案資訊: case_id={case_id}, case_name={case_name}, filename={filename}")

        # 嘗試從 Excel B2 讀取案場名稱
        project_name = None
        try:
            from openpyxl import load_workbook as _lw
            _wb = _lw(file_path, data_only=True, read_only=True)
            _ws = _wb.active
            _val = _ws['B2'].value
            if _val and str(_val).strip() and str(_val).strip() != '-':
                project_name = str(_val).strip()
            _wb.close()
        except Exception:
            pass

        log_action(session.get('user_email', 'anonymous'), 'upload_excel',
                   f"case={case_name}({case_id}) file={original_filename}")
        return jsonify({
            "status": "success",
            "message": f"檔案已上傳: {original_filename}",
            "file_path": file_path,
            "filename": filename,
            "original_filename": original_filename,
            "project_name": project_name
        })
    except Exception as e:
        import traceback
        print(f"檔案上傳錯誤: {e}")
        print(traceback.format_exc())
        log_error(session.get('user_email', 'anonymous'), 'upload_excel', e)
        return jsonify({"error": f"檔案上傳失敗: {str(e)}"}), 500


@agent_bp.route('/read_excel/<case_id>', methods=['GET'])
def read_excel(case_id):
    """
    讀取指定案場的 Excel 檔案並轉換為前端格式
    """
    try:
        excel_dir = _get_excel_dir()

        # 從查詢參數獲取前端名稱和原始檔名
        case_name = request.args.get('case_name', '')
        original_filename = request.args.get('original_filename', '')

        print(f"讀取 Excel 請求: case_id={case_id}, case_name={case_name}, original_filename={original_filename}")

        excel_file = None

        # 如果有前端名稱和原始檔名，直接組成完整檔名
        if case_name and original_filename:
            full_filename = f"{case_name}_{original_filename}"
            file_path = os.path.join(excel_dir, full_filename)
            if os.path.exists(file_path):
                excel_file = file_path
                print(f"找到檔案: {full_filename}")

        # 如果找不到，嘗試搜尋所有符合的檔案
        if not excel_file and os.path.exists(excel_dir):
            existing_files = os.listdir(excel_dir)
            print(f"搜尋 Excel 目錄中的檔案: {existing_files}")
            for filename in existing_files:
                if filename.endswith(('.xlsx', '.xls')):
                    excel_file = os.path.join(excel_dir, filename)
                    print(f"使用找到的第一個 Excel 檔案: {filename}")
                    break

        if not excel_file or not os.path.exists(excel_file):
            return jsonify({"error": "找不到指定案場的 Excel 檔案"}), 404

        # 使用 openpyxl 讀取 Excel（data_only=True 取計算後的值）
        import openpyxl

        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheets_data = []

        for sheet_index, sheet_name in enumerate(wb.sheetnames):
            sheet = wb[sheet_name]
            celldata = []

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None or cell.value == '':
                        continue

                    v_obj = {
                        'v': cell.value,
                        'm': str(cell.value),
                        'ct': {
                            'fa': 'General',
                            't': 'n' if isinstance(cell.value, (int, float)) else 's'
                        }
                    }
                    celldata.append({
                        'r': cell.row - 1,
                        'c': cell.column - 1,
                        'v': v_obj
                    })

            max_row = sheet.max_row    if sheet.max_row    else 20
            max_col = sheet.max_column if sheet.max_column else 10

            sheets_data.append({
                'name': sheet_name,
                'index': str(sheet_index),
                'status': '1' if sheet_index == 0 else '0',
                'order': str(sheet_index),
                'celldata': celldata,
                'row': max(max_row, 20),
                'column': max(max_col, 10),
                'config': {},
                'pivotTable': None,
                'isPivotTable': False,
                'luckysheet_select_save': [],
                'calcChain': [],
                'hyperlink': {},
                'dataVerification': {}
            })

        print(f"成功讀取 Excel: {excel_file}, 共 {len(sheets_data)} 個工作表")
        log_action(session.get('user_email', 'anonymous'), 'read_excel',
                   f"case_id={case_id} file={os.path.basename(excel_file)} sheets={len(sheets_data)}")
        return jsonify({
            'status': 'success',
            'data': sheets_data,
            'filename': os.path.basename(excel_file)
        })

    except Exception as e:
        import traceback
        print(f"讀取 Excel 錯誤: {e}")
        print(traceback.format_exc())
        log_error(session.get('user_email', 'anonymous'), 'read_excel', e, f"case_id={case_id}")
        return jsonify({"error": f"讀取 Excel 失敗: {str(e)}"}), 500

@agent_bp.route('/rename_excel', methods=['POST'])
def rename_excel():
    """
    重新命名後端 Excel 檔案
    """
    try:
        data = request.json
        case_id = data.get('case_id')
        old_filename = data.get('old_filename')  # 完整檔名（例如：表 1_公版.xlsx）
        new_filename = data.get('new_filename')  # 完整檔名（例如：財報_公版.xlsx）

        print(f"重新命名請求: case_id={case_id}, old={old_filename}, new={new_filename}")

        if not all([old_filename, new_filename]):
            return jsonify({"error": "缺少必要參數"}), 400

        excel_dir = _get_excel_dir()

        # 直接使用完整檔名（不再加上 case_id 前綴）
        old_path = os.path.join(excel_dir, old_filename)
        new_path = os.path.join(excel_dir, new_filename)

        print(f"檢查檔案: {old_path}")
        print(f"目標路徑: {new_path}")

        # 檢查舊檔案是否存在
        if not os.path.exists(old_path):
            # 列出目錄中的所有檔案幫助除錯
            if os.path.exists(excel_dir):
                existing_files = os.listdir(excel_dir)
                print(f"找不到檔案，Excel 目錄中的檔案: {existing_files}")
            return jsonify({"error": f"找不到原始檔案: {old_filename}"}), 404

        # 重新命名檔案
        os.rename(old_path, new_path)

        print(f"檔案已重新命名: {old_filename} -> {new_filename}")
        log_action(session.get('user_email', 'anonymous'), 'rename_excel',
                   f"old={old_filename} new={new_filename}")
        return jsonify({
            "status": "success",
            "message": "檔案重新命名成功",
            "new_filename": new_filename
        })

    except Exception as e:
        import traceback
        print(f"🚨 重新命名錯誤: {e}")
        print(traceback.format_exc())
        log_error(session.get('user_email', 'anonymous'), 'rename_excel', e,
                  f"old={old_filename if 'old_filename' in dir() else ''}")
        return jsonify({"error": f"重新命名失敗: {str(e)}"}), 500

@agent_bp.route('/delete_excel', methods=['POST'])
def delete_excel():
    """
    刪除後端 Excel 檔案
    """
    try:
        data = request.json
        case_id = data.get('case_id')
        filename = data.get('filename')  # 完整檔名（例如：財報_公版.xlsx）

        print(f"刪除 Excel 請求: case_id={case_id}, filename={filename}")

        if not filename:
            return jsonify({"error": "缺少檔案名稱參數"}), 400

        excel_dir = _get_excel_dir()
        file_path = os.path.join(excel_dir, filename)

        print(f"檢查檔案: {file_path}")

        # 檢查檔案是否存在
        if not os.path.exists(file_path):
            # 列出目錄中的所有檔案幫助除錯
            if os.path.exists(excel_dir):
                existing_files = os.listdir(excel_dir)
                print(f"找不到檔案，Excel 目錄中的檔案: {existing_files}")
            return jsonify({"error": f"找不到檔案: {filename}"}), 404

        # 刪除檔案
        os.remove(file_path)

        print(f"檔案已刪除: {filename}")
        log_action(session.get('user_email', 'anonymous'), 'delete_excel',
                   f"file={filename}")

        return jsonify({
            "status": "success",
            "message": "檔案刪除成功",
            "deleted_filename": filename
        })

    except Exception as e:
        import traceback
        print(f"🚨 刪除檔案錯誤: {e}")
        print(traceback.format_exc())
        log_error(session.get('user_email', 'anonymous'), 'delete_excel', e,
                  f"file={filename if 'filename' in dir() else ''}")
        return jsonify({"error": f"刪除檔案失敗: {str(e)}"}), 500


@agent_bp.route('/save_excel', methods=['POST'])
def save_excel():
    """
    將 Luckysheet 數據保存回 Excel 檔案（保留公式）

    重要：此函數會保護原始 Excel 中的公式不被覆蓋。
    如果原始儲存格有公式，且 Luckysheet 數據中沒有公式資訊，則保留原始公式。
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        data = request.json
        case_name = data.get('case_name', '')
        sheets_data = data.get('sheets', [])

        if not case_name:
            return jsonify({"status": "error", "error": "缺少案場名稱"}), 400

        excel_dir = _get_excel_dir()
        excel_path = None

        # 搜尋對應的 Excel 檔案
        if os.path.exists(excel_dir):
            for f in os.listdir(excel_dir):
                if f.startswith(f"{case_name}_") and f.endswith(('.xlsx', '.xls')) and not f.startswith('~$'):
                    excel_path = os.path.join(excel_dir, f)
                    break

        if not excel_path or not os.path.exists(excel_path):
            return jsonify({"status": "error", "error": f"找不到案場「{case_name}」的 Excel 檔案"}), 404

        print(f"[save_excel] 保存到: {excel_path}")

        # 讀取原始 Excel（保留公式）
        workbook = openpyxl.load_workbook(excel_path, data_only=False)

        # 建立 sheet 名稱到索引的映射
        sheet_name_map = {sheet.get('name'): sheet for sheet in sheets_data}

        formulas_preserved = 0  # 統計保留的公式數量

        for sheet_name in workbook.sheetnames:
            if sheet_name not in sheet_name_map:
                continue

            ws = workbook[sheet_name]
            sheet_data = sheet_name_map[sheet_name]
            celldata = sheet_data.get('celldata', [])

            # 建立座標到儲存格數據的映射
            cell_map = {}
            for cell in celldata:
                r = cell.get('r', 0)
                c = cell.get('c', 0)
                cell_map[(r, c)] = cell

            # 更新儲存格
            for (r, c), cell_info in cell_map.items():
                v = cell_info.get('v', {})
                excel_cell = ws.cell(row=r+1, column=c+1)

                # 檢查原始儲存格是否有公式
                original_has_formula = False
                if excel_cell.value and isinstance(excel_cell.value, str) and excel_cell.value.startswith('='):
                    original_has_formula = True

                if isinstance(v, dict):
                    # 如果 Luckysheet 數據有公式，使用公式
                    if 'f' in v and v['f']:
                        excel_cell.value = '=' + v['f'] if not v['f'].startswith('=') else v['f']
                    elif original_has_formula:
                        # 原始儲存格有公式，但 Luckysheet 沒有公式資訊，保留原始公式
                        formulas_preserved += 1
                        # 不做任何修改，保留原始公式
                    elif 'v' in v:
                        # 原始沒有公式，使用 Luckysheet 的值
                        excel_cell.value = v['v']
                else:
                    if original_has_formula:
                        # 保留原始公式
                        formulas_preserved += 1
                    else:
                        excel_cell.value = v

        # 保存檔案
        workbook.save(excel_path)
        workbook.close()

        print(f"[save_excel] 保存成功，保留了 {formulas_preserved} 個公式")
        log_action(session.get('user_email', 'anonymous'), 'save_excel',
                   f"case={case_name} file={os.path.basename(excel_path)} formulas_preserved={formulas_preserved}")
        return jsonify({"status": "success", "message": f"Excel 已保存（保留了 {formulas_preserved} 個公式）"})

    except Exception as e:
        import traceback
        print(f"[save_excel] 保存失敗: {e}")
        print(traceback.format_exc())
        log_error(session.get('user_email', 'anonymous'), 'save_excel', e,
                  f"case={case_name if 'case_name' in dir() else ''}")
        return jsonify({"status": "error", "error": f"保存失敗: {str(e)}"}), 500


@agent_bp.route('/delete_sheet', methods=['POST'])
def delete_sheet():
    """
    從後端 Excel 檔案中刪除指定工作表（前端手動刪除觸發）
    Body: { case_id, case_name, original_filename, sheet_name }
    """
    try:
        data = request.json
        case_id = data.get('case_id', '')
        case_name = data.get('case_name', '')
        original_filename = data.get('original_filename', '')
        sheet_name = data.get('sheet_name', '').strip()

        user_email = session.get('user_email', 'anonymous')

        if not sheet_name:
            return jsonify({"error": "缺少工作表名稱"}), 400

        excel_path = _find_excel_file(
            case_id=case_id,
            case_name=case_name,
            original_filename=original_filename
        )
        if not excel_path:
            return jsonify({"error": "找不到對應的 Excel 檔案"}), 404

        from tool.tool_manager import ToolManager
        tm = ToolManager()
        tm.set_excel_file(excel_path)
        result = tm.execute_tool("delete_excel_sheet", {"sheet_name": sheet_name})

        if result.get("success"):
            log_action(user_email, 'delete_sheet',
                       f"case={case_name}({case_id}) sheet={sheet_name}")
            return jsonify({
                "status": "success",
                "message": result.get("message", ""),
                "deleted_sheet": result.get("deleted_sheet", sheet_name),
                "remaining_sheets": result.get("remaining_sheets", [])
            })
        else:
            return jsonify({"error": result.get("message", "刪除失敗")}), 400

    except Exception as e:
        import traceback
        print(f"🚨 刪除工作表錯誤: {e}")
        print(traceback.format_exc())
        log_error(session.get('user_email', 'anonymous'), 'delete_sheet', e)
        return jsonify({"error": f"刪除工作表失敗: {str(e)}"}), 500
